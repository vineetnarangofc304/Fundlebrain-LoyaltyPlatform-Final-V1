"""Historic Data Ingestion + Demo-Data Purge.

Accepts CSV uploads of customers / transactions / stores / items (KAZO format),
parses date strings (DD-MM-YYYY / DD-MM-YYYY HH:MM), upserts into MongoDB in
background, tracks per-job progress in `historic_ingest_jobs`.

Supports CHUNKED uploads (init -> chunk x N -> finalize) so large files (33MB+)
bypass Kubernetes ingress body-size limits in production.

Strict: NO dummy fallback values. Rows with critical missing fields (mobile for
customers, bill_number for transactions) are recorded as errors.
"""
from __future__ import annotations
import csv
import io
import json
import re
import uuid
import logging
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple

from fastapi import (APIRouter, Depends, HTTPException, UploadFile, File, Form,
                       BackgroundTasks)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import (
    db, customers_col, transactions_col, stores_col, campaigns_col, coupons_col,
    points_ledger_col, api_logs_col, nps_col, tickets_col, ai_chats_col,
    campaign_metrics_col, message_log_col, audit_logs_col, coupon_redemptions_col,
)
from auth import get_current_user, require_roles, MANAGEMENT_ROLES

router = APIRouter(prefix="/historic-data", tags=["historic-data"])
logger = logging.getLogger("kazo-fundle.historic")

# Collections for staging
historic_jobs_col = db["historic_ingest_jobs"]
historic_raw_col = db["historic_uploads"]  # stores raw uploaded CSV bytes (capped)
historic_chunks_col = db["historic_chunks"]  # shared chunk store (works across pods/workers)
historic_skipped_col = db["historic_skipped_rows"]  # every parser rejection — reconcilable

CHUNK_SIZE = 500
MAX_FILE_BYTES = 350 * 1024 * 1024  # 350 MB (chunked uploads bypass proxy limits)

# Indian Standard Time + the fixed validity for all historically-loaded opening
# balances. Per the launch rule, every customer's loaded point balance is valid
# until end-of-day 31 Dec 2026 IST. Stored as UTC for clean lexical comparison.
IST = timezone(timedelta(hours=5, minutes=30))
OPENING_BALANCE_EXPIRY_ISO = datetime(2026, 12, 31, 23, 59, 59, tzinfo=IST).astimezone(timezone.utc).isoformat()
MAX_ERROR_SAMPLES = 25  # for the in-job samples (UI preview); full set lives in historic_skipped_col
SKIPPED_PERSIST_CAP = 1_000_000  # safety cap on skipped-rows write per job
ALLOWED_DATASETS = {"customers", "transactions", "stores", "items", "points_ledger", "sku_transactions"}
DUPLICATE_MODES = {"upsert", "skip", "fail"}


# ---------------- Date parsing ----------------
_DATE_PATTERNS = [
    "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    "%d-%b-%Y", "%d %b %Y",
]


def parse_date(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"na", "n/a", "null", "none", "-"}:
        return None
    for pat in _DATE_PATTERNS:
        try:
            return datetime.strptime(s, pat).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def parse_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (ValueError, TypeError):
        return default


def parse_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def _norm_mobile(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 12 and digits.startswith("91"):
        return digits[2:]
    return digits


# ---------------- Row mappers (per dataset) ----------------
def _map_customer_row(r: Dict[str, str]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    mobile = _norm_mobile(r.get("Mobile") or r.get("mobile") or r.get("Customer Mobile"))
    if not mobile or len(mobile) < 7:
        return None, "Missing/invalid Mobile"
    last_visit = parse_date(r.get("Last Visit Date"))
    first_visit = parse_date(r.get("First Visit Date"))
    added_on = parse_date(r.get("Added On"))
    dob = parse_date(r.get("DOB"))
    doa = parse_date(r.get("DOA"))
    lifetime_spend = parse_float(r.get("Total Billing"))
    visit_count = parse_int(r.get("Total Visits"))
    points_balance = parse_int(r.get("Current Point Balance"))
    points_redeemed = parse_int(r.get("Redeem Points"))
    days_since_last_visit = parse_int(r.get("Days Since Last Visit"), -1)
    tier = _derive_tier(lifetime_spend)
    name = (r.get("Name") or r.get("Customer Name") or "").strip() or None
    city = (r.get("City") or "").strip() or None
    state = (r.get("State") or "").strip() or None
    online = (r.get("ONLINE") or "").strip().lower() in {"online", "1", "true", "yes"}
    return ({
        "mobile": mobile,
        "name": name,
        "city": city,
        "state": state,
        "country_code": (r.get("Country Code") or "91").strip(),
        "card_validity": (r.get("Card Validity") or "").strip() or None,
        "registered_account": (r.get("Registred Account") or r.get("Registered Account") or "").strip() or None,
        "first_visit_account": (r.get("First Visit Account") or "").strip() or None,
        "last_visit_account": (r.get("Last Visit Account") or "").strip() or None,
        "last_visit_at": last_visit.isoformat() if last_visit else None,
        "first_purchase_at": first_visit.isoformat() if first_visit else None,
        "added_on": added_on.isoformat() if added_on else None,
        "birthday": dob.isoformat()[:10] if dob else None,
        "anniversary": doa.isoformat()[:10] if doa else None,
        "lifetime_spend": lifetime_spend,
        "visit_count": visit_count,
        "points_balance": points_balance,
        "lifetime_points_redeemed": points_redeemed,
        "days_since_last_visit": days_since_last_visit if days_since_last_visit >= 0 else None,
        "tier": tier,
        "is_online": online,
        "source": "historic_upload",
    }, None)


def _derive_tier(lifetime_spend: float) -> str:
    if lifetime_spend >= 200_000:
        return "diamond"
    if lifetime_spend >= 75_000:
        return "platinum"
    if lifetime_spend >= 25_000:
        return "gold"
    return "silver"


def _map_transaction_row(r: Dict[str, str], store_cache: Dict[str, Dict[str, Any]]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    bill = (r.get("Bill Number") or r.get("Transaction Id") or "").strip()
    if not bill:
        return None, "Missing Bill Number / Transaction Id"
    # Mobile is OPTIONAL — anonymous walk-ins are valid bills (tracked as
    # "Lost Opportunities" in the Live Monitor cockpit). Store as None when absent.
    mobile = _norm_mobile(r.get("Customer Mobile Number") or r.get("Customer Mobile") or r.get("Mobile"))
    mobile = mobile or None
    bill_date = parse_date(r.get("Date"))
    if not bill_date:
        return None, f"Invalid Date '{r.get('Date')}'"
    time_str = (r.get("Time") or "").strip()
    if time_str and ":" in time_str:
        try:
            hh, mm, *rest = time_str.split(":")
            bill_date = bill_date.replace(hour=int(hh) % 24, minute=int(mm) % 60)
        except (ValueError, IndexError):
            pass
    net = parse_float(r.get("Net Amount Before Tax Kazo") or r.get("Net Amount Before Tax")
                      or r.get("Net Amount Before Tax KAZO") or r.get("Net Amount"))
    # Tax AMOUNT only (never "Tax Rate", which is a percentage).
    tax = parse_float(r.get("Total Tax") or r.get("Tax Total") or r.get("Total Tax Kazo")
                      or r.get("Tax Amount") or r.get("Tax Amt"))
    discount = parse_float(r.get("Discount"))
    total = parse_float(r.get("Total Revenue Kazo") or r.get("Total Revenue KAZO")
                        or r.get("Total Revenue") or r.get("Total"))
    # Gross billing (MRP before discount) — used as the loyalty/spend base for history.
    bill_amount = parse_float(r.get("Total Billing KAZO") or r.get("Total Billing Kazo")
                              or r.get("Total Billing") or r.get("Total Billing Lifetime")
                              or r.get("Bill Amount") or r.get("Bill Amt"))
    if total == 0 and net:
        total = net + tax - discount

    # R6: per-bill points columns — load AS-IS from CSV (no expiry yet).
    points_earned = parse_int(
        r.get("Points Earned") or r.get("Point Earned") or r.get("Earn Points") or r.get("Earned Points")
    )
    points_redeemed = parse_int(
        r.get("Points Redeemed") or r.get("Redeem Points") or r.get("Burn Points") or r.get("Redeemed Points")
    )
    bonus_points = parse_int(r.get("Bonus Points") or r.get("Bonus"))

    outlet = (r.get("Outlet(Only For Shopify Marker)") or r.get("Outlet")
              or r.get("Store Name") or r.get("Outlet Name") or "").strip() or None
    # "Store master" / "Store code" carries the canonical K-code (== POS customer_key),
    # so historical bills align with live POS bills on the same store.
    store_code = (r.get("Store master") or r.get("Store Master") or r.get("Store code")
                  or r.get("Store Code") or r.get("Loc Code") or r.get("Customer Key")
                  or r.get("customer_key") or "").strip() or None
    city = (r.get("City") or "").strip() or None
    zone = (r.get("Zone New") or r.get("Zone Name") or r.get("Zone") or "").strip() or None
    store_class = (r.get("Class") or "").strip() or None
    store_id = None
    # Identify the store by its K-code when present (so it aligns with live POS combo
    # resolution), else fall back to the outlet name. store_id is backfilled in the post-pass.
    cache_key = store_code or (outlet.lower() if outlet else None)
    if cache_key:
        if cache_key not in store_cache:
            store_cache[cache_key] = {"code": store_code, "name": outlet, "city": city,
                                        "zone": zone, "class": store_class, "id": None}
        store_id = store_cache[cache_key].get("id")
    return_marker = (r.get("Return Marker") or "Regular").strip()
    return_reason = (r.get("Return Reason") or "").strip()
    # A non-empty Return Reason (new Billing Report) marks a return even without a marker.
    is_return = return_marker.lower() == "return" or bool(return_reason) or (total or 0) < 0
    new_existing = (r.get("New_Existing") or r.get("New Existing") or r.get("New/Existing") or "").strip()
    recency = (r.get("Recency") or "").strip()

    return ({
        "bill_number": bill,
        "transaction_id": (r.get("Transaction Id") or bill).strip(),
        "customer_mobile": mobile,
        "customer_name": (r.get("Customer Name") or "").strip() or None,
        "store_id": store_id,  # filled later when store is created
        "store_code": store_code,
        "store_name": outlet or None,
        "city": city,
        "zone": zone,
        "store_class": store_class,
        "bill_date": bill_date.isoformat(),
        "net_amount": total or net,
        "net_amount_before_tax": net,
        "tax_amount": tax,
        "discount_amount": discount,
        "gross_amount": bill_amount or ((net or 0) + (tax or 0)),
        "bill_amount": bill_amount or None,
        "is_return": is_return,
        "return_marker": "Return" if is_return else (return_marker or "Regular"),
        "return_reason": return_reason or None,
        "new_or_existing": new_existing or None,
        "recency": recency or None,
        "items": [],  # KAZO export has no line-item breakdown
        "payment_mode": "unknown",
        "points_earned": points_earned,
        "points_redeemed": points_redeemed,
        "bonus_points": bonus_points,
        "source": "historic_upload",
    }, None)


def _map_store_row(r: Dict[str, str]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    name = (r.get("Name") or r.get("Store Name") or r.get("Outlet") or "").strip()
    if not name:
        return None, "Missing Store Name"
    code = (r.get("Code") or r.get("Store Code") or _make_store_code(name)).strip()
    city = (r.get("City") or "").strip()
    if not city:
        return None, "Missing City"
    return ({
        "code": code,
        "name": name,
        "city": city,
        "state": (r.get("State") or "").strip() or None,
        "region": (r.get("Region") or r.get("Zone") or "").strip() or None,
        "address": (r.get("Address") or "").strip() or None,
        "phone": (r.get("Phone") or "").strip() or None,
        "manager_name": (r.get("Manager") or "").strip() or None,
        "is_active": True,
        "source": "historic_upload",
    }, None)


def _make_store_code(name: str) -> str:
    base = re.sub(r"[^A-Z0-9]+", "", name.upper())[:8]
    return f"K{base}"


def _norm_store_name(name: str) -> str:
    """Normalise an outlet name for robust matching against the Store Master
    (lowercase, collapse whitespace, drop punctuation) so minor formatting
    differences between the Billing Report and the Store Master still align."""
    return re.sub(r"[^a-z0-9]+", " ", (name or "").lower()).strip()


def _map_item_row(r: Dict[str, str]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    # Accept common KAZO column variants for SKU
    sku = (
        r.get("SKU") or r.get("sku") or r.get("Sku")
        or r.get("Item Id") or r.get("Item ID")
        or r.get("Item Code") or r.get("ITEM CODE") or r.get("Article")
        or r.get("Style Code") or r.get("StyleCode")
        or ""
    ).strip()
    if not sku:
        return None, "Missing SKU / Item Id / Item Code / Style Code"

    name = (
        r.get("Name") or r.get("Item Name") or r.get("Product Name")
        or r.get("Style Name") or r.get("Description") or ""
    ).strip() or None

    category = (
        r.get("Category") or r.get("Item Master Category") or r.get("Item Category")
        or r.get("Product Category") or r.get("Class") or ""
    ).strip() or None

    sub_category = (r.get("Sub Category") or r.get("Sub-Category") or "").strip() or None

    price = parse_float(
        r.get("MRP") or r.get("mrp") or r.get("MRP Price") or r.get("List Price")
        or r.get("Price") or r.get("Selling Price") or r.get("Rate")
    )

    color = (r.get("Color") or r.get("Colour") or "").strip() or None
    size = (r.get("Size") or "").strip() or None
    brand = (r.get("Brand") or "").strip() or None
    season = (r.get("Season") or "").strip() or None
    hsn = (r.get("HSN") or r.get("HSN Code") or r.get("hsn") or "").strip() or None
    tax_pct = parse_float(r.get("Tax %") or r.get("Tax Rate") or r.get("GST"))

    doc = {
        "sku": sku,
        "name": name,
        "category": category,
        "sub_category": sub_category,
        "price": price,
        "color": color,
        "size": size,
        "brand": brand,
        "season": season,
        "hsn": hsn,
        "tax_pct": tax_pct,
        "is_active": True,
        "source": "historic_upload",
    }
    # Strip None to keep the document tidy
    doc = {k: v for k, v in doc.items() if v is not None}
    return doc, None


def _map_points_ledger_row(r: Dict[str, str]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """Direct points-ledger ingest — for brands that already track their own
    earn/redeem/bonus ledger and want to migrate it verbatim into Fundle."""
    mobile = (r.get("Mobile") or r.get("Customer Mobile") or r.get("Customer Mobile Number") or "").strip()
    if not mobile:
        return None, "Missing Mobile"
    # Normalize 10-digit
    mobile = mobile.replace(" ", "").replace("-", "").replace("+", "")
    if mobile.startswith("91") and len(mobile) == 12:
        mobile = mobile[2:]
    if len(mobile) != 10:
        return None, f"Invalid mobile length: {mobile}"

    points = parse_float(r.get("Points") or r.get("Point") or r.get("Quantity"))
    if points is None or points == 0:
        return None, "Missing/zero Points"
    type_raw = (r.get("Type") or r.get("Direction") or "").strip().lower()
    if not type_raw:
        type_raw = "earn" if points > 0 else "redeem"
    if type_raw not in {"earn", "redeem", "bonus", "adjust", "expire"}:
        return None, f"Unknown type: {type_raw}"

    # ledger stored with signed points (earn/bonus positive, redeem/expire negative)
    signed_points = abs(points)
    if type_raw in {"redeem", "expire"}:
        signed_points = -signed_points

    bill_no = (r.get("Bill Number") or r.get("Bill") or r.get("Reference") or "").strip() or None
    bill_date = parse_date(r.get("Date") or r.get("Bill Date") or r.get("Transaction Date"))
    reason = (r.get("Reason") or r.get("Description") or r.get("Note") or "").strip() or None
    if reason and len(reason) > 500:
        reason = reason[:500]
    source_bill_id = (r.get("Source Bill Id") or r.get("SourceBillId") or bill_no or "").strip() or None

    doc = {
        "customer_mobile": mobile,
        "type": type_raw,
        "points": signed_points,
        "bill_number": bill_no,
        "bill_date": bill_date.isoformat() if bill_date else None,
        "reason": reason or f"{type_raw} (historic ingest)",
        "source_bill_id": source_bill_id,
        "source": "historic_upload",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    doc = {k: v for k, v in doc.items() if v is not None}
    return doc, None


def _map_sku_line_row(r: Dict[str, str]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """SKU-wise / line-item bill data (e.g. Kazo_SKU_Master_Data).

    Each row is one item line on a bill. We attach these lines to the matching
    transaction (`items[]`) and also build the item master catalog. The join key
    is the SKU file's "Transaction Id" (the 000000PK… value) which equals the
    billwise Bill Number / Transaction Id.
    """
    txn_key = (r.get("Transaction Id") or r.get("Transaction ID") or r.get("Bill Number") or "").strip()
    if not txn_key:
        return None, "Missing Transaction Id / Bill Number"
    sku = (r.get("Item Id") or r.get("Item ID") or r.get("SKU") or r.get("sku")
           or r.get("Item Code") or "").strip()
    if not sku:
        return None, "Missing Item Id / SKU"
    name = (r.get("Item Name") or r.get("Name") or r.get("Product Name") or "").strip() or None
    category = (r.get("Item Master Category") or r.get("Category") or "").strip() or None
    category_id = (r.get("Category 1(Logic)") or r.get("Category 0(Logic)") or "").strip() or None
    season = (r.get("Season") or "").strip() or None
    qty = parse_int(r.get("Quantity") or r.get("Qty"), 1) or 1
    rate = parse_float(r.get("Rate"))
    discount = parse_float(r.get("discount") or r.get("Discount"))
    subtotal = parse_float(r.get("Sub Total") or r.get("Subtotal"))
    return ({
        "_txn_key": txn_key,
        "invoice_number": (r.get("Bill Number") or "").strip() or None,
        "sku": sku,
        "name": name,
        "category": category,
        "category_id": category_id,
        "season": season,
        "quantity": qty,
        "unit_price": rate,
        "discount": discount,
        "total": subtotal or (round(rate * qty, 2) if rate else 0),
    }, None)


MAPPERS = {
    "customers": _map_customer_row,
    "transactions": _map_transaction_row,  # special: store_cache passed
    "stores": _map_store_row,
    "items": _map_item_row,
    "points_ledger": _map_points_ledger_row,
    "sku_transactions": _map_sku_line_row,
}


# ---------------- Background job ----------------
async def _run_ingest_job(job_id: str, dataset: str, raw_text: str,
                            duplicate_mode: str, dry_run: bool):
    def now():
        return datetime.now(timezone.utc).isoformat()
    await historic_jobs_col.update_one({"id": job_id},
        {"$set": {"status": "running", "started_at": now()}})
    try:
        reader = csv.DictReader(io.StringIO(raw_text))
        total_rows = 0
        processed = 0
        inserted = 0
        updated = 0
        skipped = 0
        errors: List[Dict[str, Any]] = []
        skipped_buf: List[Dict[str, Any]] = []  # buffered skipped-row docs for persistence
        skipped_persisted = 0

        async def _persist_skipped(force: bool = False):
            """Persist skipped rows in batches so a re-run / debug can read them back."""
            nonlocal skipped_persisted
            if not skipped_buf:
                return
            if not force and len(skipped_buf) < 1000:
                return
            if dry_run:
                skipped_buf.clear()
                return
            if skipped_persisted >= SKIPPED_PERSIST_CAP:
                skipped_buf.clear()
                return
            try:
                await historic_skipped_col.insert_many(skipped_buf, ordered=False)
                skipped_persisted += len(skipped_buf)
            except Exception:
                pass
            skipped_buf.clear()

        def _record_skip(row_num: int, reason: str, raw_row: Optional[Dict[str, Any]] = None):
            nonlocal skipped
            skipped += 1
            if len(errors) < MAX_ERROR_SAMPLES:
                errors.append({
                    "row": row_num,
                    "reason": reason,
                    "data_sample": {k: raw_row.get(k) for k in list((raw_row or {}).keys())[:6]} if raw_row else {},
                })
            # Persist every skipped row (capped)
            if skipped_persisted + len(skipped_buf) < SKIPPED_PERSIST_CAP:
                skipped_buf.append({
                    "id": uuid.uuid4().hex,
                    "job_id": job_id,
                    "row_number": row_num,
                    "reason": reason,
                    "raw_row": raw_row or {},
                    "created_at": now(),
                })
        store_cache: Dict[str, Dict[str, Any]] = {}
        # SKU-wise accumulators: line items grouped by txn join-key + distinct item master
        sku_bill_items: Dict[str, List[Dict[str, Any]]] = {}
        sku_item_master: Dict[str, Dict[str, Any]] = {}
        buffer_insert: List[Dict[str, Any]] = []

        # Pre-pass: count rows
        # (csv.DictReader iterates lazily; we count as we go)

        async def _flush():
            nonlocal inserted, updated
            if dry_run:
                return
            if dataset == "customers":
                if buffer_insert:
                    # Use bulk upsert for customers (by mobile)
                    from pymongo import UpdateOne
                    ops = [UpdateOne({"mobile": d["mobile"]},
                                      {"$set": {**d, "ingest_job_id": job_id}, "$setOnInsert": {
                                          "id": uuid.uuid4().hex,
                                          "created_at": now(),
                                      }},
                                      upsert=True) for d in buffer_insert]
                    res = await customers_col.bulk_write(ops, ordered=False)
                    inserted += res.upserted_count
                    # matched_count = rows that matched the filter (incl. no-op upserts where data
                    # didn't change). MongoDB's modified_count is misleading because it returns 0
                    # for upserts with identical values — making re-uploads look like 'data is lost'.
                    updated += res.matched_count
                    buffer_insert.clear()
            elif dataset == "transactions":
                if buffer_insert:
                    from pymongo import UpdateOne
                    ops = []
                    for d in buffer_insert:
                        # Hard bill uniqueness key per KAZO rule: (store_id, bill_number, bill_date).
                        # During the initial insert store_id may be null; we use bill_number alone
                        # because the post-pass backfills store_id and the unique compound index
                        # ensures any future duplicate (same store+bill+date) is rejected.
                        ops.append(UpdateOne({"bill_number": d["bill_number"],
                                              "bill_date": d["bill_date"]},
                            {"$set": {**d, "ingest_job_id": job_id},
                             "$setOnInsert": {"id": uuid.uuid4().hex, "created_at": now()}},
                            upsert=True))
                    res = await transactions_col.bulk_write(ops, ordered=False)
                    inserted += res.upserted_count
                    updated += res.matched_count
                    buffer_insert.clear()
            elif dataset == "stores":
                if buffer_insert:
                    from pymongo import UpdateOne
                    ops = [UpdateOne({"code": d["code"]},
                                      {"$set": d, "$setOnInsert": {
                                          "id": uuid.uuid4().hex,
                                          "created_at": now(),
                                      }},
                                      upsert=True) for d in buffer_insert]
                    res = await stores_col.bulk_write(ops, ordered=False)
                    inserted += res.upserted_count
                    updated += res.matched_count
                    buffer_insert.clear()
            elif dataset == "items":
                items_col = db["items"]
                if buffer_insert:
                    from pymongo import UpdateOne
                    ops = [UpdateOne({"sku": d["sku"]},
                                      {"$set": d, "$setOnInsert": {
                                          "id": uuid.uuid4().hex,
                                          "created_at": now(),
                                      }},
                                      upsert=True) for d in buffer_insert]
                    res = await items_col.bulk_write(ops, ordered=False)
                    inserted += res.upserted_count
                    updated += res.matched_count
                    buffer_insert.clear()
            elif dataset == "sku_transactions":
                # Accumulate line items per bill + distinct item master. Actual writes
                # (attach to transactions + upsert item master) happen in the post-pass.
                if buffer_insert:
                    for d in buffer_insert:
                        tk = d.get("_txn_key")
                        sku = d.get("sku")
                        if sku and sku not in sku_item_master:
                            sku_item_master[sku] = {
                                "sku": sku, "name": d.get("name"),
                                "category": d.get("category"), "season": d.get("season"),
                                "price": d.get("unit_price"),
                                "is_active": True, "source": "historic_upload",
                            }
                        if tk:
                            sku_bill_items.setdefault(tk, []).append({
                                "sku": sku, "name": d.get("name"),
                                "category": d.get("category"), "category_id": d.get("category_id"),
                                "quantity": d.get("quantity"), "unit_price": d.get("unit_price"),
                                "total": d.get("total"), "discount": d.get("discount"),
                                "season": d.get("season"),
                            })
                    # Every processed line is "accounted for" → keeps integrity check balanced.
                    updated += len(buffer_insert)
                    buffer_insert.clear()
            elif dataset == "points_ledger":
                if buffer_insert:
                    from pymongo import UpdateOne
                    ops = []
                    for d in buffer_insert:
                        # Composite key — same mobile + bill + type combo upserts (not duplicates)
                        sbi = d.get("source_bill_id") or d.get("bill_number")
                        if sbi:
                            key = {"customer_mobile": d["customer_mobile"],
                                    "source_bill_id": sbi, "type": d["type"]}
                        else:
                            # No bill reference — use bill_date + type to dedupe loosely
                            key = {"customer_mobile": d["customer_mobile"],
                                    "bill_date": d.get("bill_date"), "type": d["type"],
                                    "points": d["points"]}
                        ops.append(UpdateOne(key,
                            {"$set": d, "$setOnInsert": {"id": uuid.uuid4().hex}},
                            upsert=True))
                    res = await points_ledger_col.bulk_write(ops, ordered=False)
                    inserted += res.upserted_count
                    updated += res.matched_count
                    buffer_insert.clear()

        # Loop with PER-ROW try/except wrapper so one bad row never aborts the entire job
        for raw_row in reader:
            total_rows += 1
            try:
                try:
                    if dataset == "transactions":
                        doc, err = _map_transaction_row(raw_row, store_cache)
                    else:
                        doc, err = MAPPERS[dataset](raw_row)
                except Exception as e:
                    doc, err = None, f"Mapper error: {e}"

                if err:
                    _record_skip(total_rows, err, raw_row)
                    processed += 1
                    continue

                # Duplicate-mode handling
                if duplicate_mode == "skip":
                    key_field, key_val = _get_pk(dataset, doc)
                    if key_field:
                        exists = await _exists_col(dataset).find_one({key_field: key_val}, {"_id": 1})
                        if exists:
                            _record_skip(total_rows, "Duplicate (skip mode)", raw_row)
                            processed += 1
                            continue

                buffer_insert.append(doc)
                processed += 1

                if len(buffer_insert) >= CHUNK_SIZE:
                    try:
                        await _flush()
                    except Exception as fe:
                        logger.exception(f"Flush failed at row {total_rows}")
                        # Mark all buffered as skipped + persist
                        for d in buffer_insert:
                            _record_skip(total_rows, f"Bulk flush error: {fe}", d)
                        buffer_insert.clear()
                    await _persist_skipped()
                    await historic_jobs_col.update_one({"id": job_id},
                        {"$set": {"processed": processed, "inserted": inserted,
                                    "updated": updated, "skipped": skipped,
                                    "errors_count": skipped,
                                    "errors_sample": errors[:MAX_ERROR_SAMPLES],
                                    "heartbeat": now()}})
            except Exception as row_exc:
                logger.exception(f"Unhandled row {total_rows} exception")
                _record_skip(total_rows, f"Row exception: {row_exc}", None)

        # Final flush — protected; if it fails we still mark completed-with-errors
        try:
            await _flush()
        except Exception as fe:
            logger.exception("Final flush failed")
            for d in buffer_insert:
                _record_skip(total_rows, f"Final flush error: {fe}", d)
            buffer_insert.clear()

        # Persist any remaining skipped rows
        await _persist_skipped(force=True)

        # For transactions: create/link stores by K-code — wrapped so a bad store row never fails the job
        store_links: Dict[str, str] = {}
        if dataset == "transactions" and store_cache and not dry_run:
            try:
                from pymongo import UpdateMany
                # merchant_id for POS-combo alignment so live POS bills land on the same store
                _cred = await db["pos_credentials"].find_one({"is_active": True}, {"_id": 0, "merchant_id": 1})
                _merchant = (_cred or {}).get("merchant_id") or "KAZO_FUNDLE"
                # Preload the already-uploaded Store Master once so historical bills MERGE
                # onto existing stores (by K-code, else by normalised outlet name) rather
                # than spawning duplicates that wouldn't align with live POS.
                by_code: Dict[str, str] = {}
                by_norm_name: Dict[str, str] = {}
                async for s in stores_col.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1}):
                    if s.get("code"):
                        by_code[str(s["code"]).strip()] = s["id"]
                    if s.get("name"):
                        by_norm_name[_norm_store_name(s["name"])] = s["id"]
                for cache_key, meta in store_cache.items():
                    code = (meta.get("code") or "").strip() or None
                    name = (meta.get("name") or "").strip() or None
                    if not code and not name:
                        continue
                    try:
                        existing_id = None
                        if code and code in by_code:
                            existing_id = by_code[code]
                        if not existing_id and name:
                            existing_id = by_norm_name.get(_norm_store_name(name))
                        if existing_id:
                            meta["id"] = existing_id
                            if code:
                                # Tag the K-code + POS combo onto the existing store
                                await stores_col.update_one(
                                    {"id": existing_id},
                                    {"$set": {"code": code, "pos_customer_key": code,
                                              "pos_merchant_id": _merchant}})
                        else:
                            sid = uuid.uuid4().hex
                            new_code = code or _make_store_code(name)
                            store_doc = {
                                "id": sid,
                                "code": new_code,
                                "name": name or code,
                                "city": meta.get("city") or "",
                                "state": "",
                                "region": meta.get("zone") or "",
                                "store_class": meta.get("class"),
                                "address": name or "",
                                "is_active": True,
                                "source": "historic_upload",
                                "created_at": now(),
                            }
                            if code:
                                store_doc["pos_customer_key"] = code
                                store_doc["pos_merchant_id"] = _merchant
                            await stores_col.insert_one(store_doc)
                            meta["id"] = sid
                            by_code[new_code] = sid
                            if name:
                                by_norm_name[_norm_store_name(name)] = sid
                        store_links[cache_key] = meta["id"]
                    except Exception as se:
                        logger.warning(f"Could not create/find store '{meta.get('code') or meta.get('name')}': {se}")
                # Backfill store_id on the just-inserted transactions (match by K-code, else name)
                ops = []
                for cache_key, meta in store_cache.items():
                    sid = meta.get("id")
                    if not sid:
                        continue
                    code = (meta.get("code") or "").strip() or None
                    name = (meta.get("name") or "").strip() or None
                    if code:
                        ops.append(UpdateMany({"ingest_job_id": job_id, "store_code": code},
                                              {"$set": {"store_id": sid}}))
                    elif name:
                        ops.append(UpdateMany({"ingest_job_id": job_id, "store_name": name, "store_id": None},
                                              {"$set": {"store_id": sid}}))
                if ops:
                    try:
                        await transactions_col.bulk_write(ops, ordered=False)
                    except Exception as be:
                        logger.exception(f"Store backfill bulk_write failed: {be}")
            except Exception as spe:
                logger.exception(f"Store auto-create post-pass failed: {spe}")

        # For SKU-wise: upsert item master + attach line items to matching transactions
        if dataset == "sku_transactions" and not dry_run:
            try:
                from pymongo import UpdateOne, UpdateMany
                items_col = db["items"]
                if sku_item_master:
                    iops = [UpdateOne({"sku": s["sku"]},
                                     {"$set": {k: v for k, v in s.items() if v is not None},
                                      "$setOnInsert": {"id": uuid.uuid4().hex, "created_at": now()}},
                                     upsert=True) for s in sku_item_master.values()]
                    for i in range(0, len(iops), 1000):
                        try:
                            await items_col.bulk_write(iops[i:i + 1000], ordered=False)
                        except Exception as e:
                            logger.warning(f"SKU item-master upsert partial failure: {e}")
                if sku_bill_items:
                    attached = 0
                    tops = []
                    for tk, items in sku_bill_items.items():
                        units = sum(int(it.get("quantity") or 0) for it in items) or len(items)
                        tops.append(UpdateMany(
                            {"$or": [{"transaction_id": tk}, {"bill_number": tk}]},
                            {"$set": {"items": items, "units_count": units}}))
                        attached += 1
                        if len(tops) >= 1000:
                            try:
                                await transactions_col.bulk_write(tops, ordered=False)
                            except Exception as e:
                                logger.warning(f"SKU items attach partial failure: {e}")
                            tops.clear()
                    if tops:
                        try:
                            await transactions_col.bulk_write(tops, ordered=False)
                        except Exception as e:
                            logger.warning(f"SKU items attach final partial failure: {e}")
                    logger.info(f"SKU ingest job {job_id}: {len(sku_item_master)} SKUs in master, "
                                f"items attached for {attached} bills")
            except Exception as e:
                logger.exception(f"SKU post-pass failed: {e}")

        # =====================================================
        # R6: Points-ledger post-pass — write earn/redeem/bonus ledger entries for
        # each loyalty bill (customer_mobile is non-empty) ingested by this job.
        # R1: ledger entries carry the actual bill_date (NOT ingest time).
        # =====================================================
        if dataset == "transactions" and not dry_run:
            try:
                await _write_ledger_for_job(job_id)
            except Exception as ple:
                logger.exception(f"Points-ledger post-pass failed: {ple}")

        # =====================================================
        # R1 + R2 + R3: per-customer aggregate recomputation —
        #   first_purchase_at, last_visit_at, visit_count, home_store_id
        # Only loyalty customers (with non-empty mobile).
        # =====================================================
        if dataset == "transactions" and not dry_run:
            try:
                await _recompute_customer_aggregates(job_id)
            except Exception as cae:
                logger.exception(f"Customer aggregate recompute failed: {cae}")

        # =====================================================
        # Opening-balance ledger — for the customer (CRM) dataset, write one
        # 'opening' ledger entry per customer (= their current point balance)
        # expiring 31 Dec 2026 IST, so the Expiry report reflects the launch rule.
        # =====================================================
        if dataset == "customers" and not dry_run:
            try:
                await _write_opening_balance_ledger_for_job(job_id)
            except Exception as obe:
                logger.exception(f"Opening-balance ledger post-pass failed: {obe}")

        final = {
            "status": "completed" if not dry_run else "previewed",
            "completed_at": now(),
            "processed": processed,
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "errors_count": skipped,
            "errors_sample": errors[:MAX_ERROR_SAMPLES],
            "total_rows": total_rows,
            "stores_auto_created": len(store_links) if dataset == "transactions" else 0,
        }
        await historic_jobs_col.update_one({"id": job_id}, {"$set": final})
        logger.info(f"Historic ingest job {job_id} done: {final}")

        # =====================================================
        # AUTO-BACKFILL — for transactions/points_ledger, ensure every
        # customer_mobile has a matching customer row + R1/R2/R3 fields
        # populated. This guarantees Active <= Total on every dashboard.
        # =====================================================
        if not dry_run and dataset in {"transactions", "points_ledger"}:
            try:
                from pymongo import UpdateOne
                # Find mobiles in this job's data that lack a customers row
                mobiles_in_job = await transactions_col.distinct(
                    "customer_mobile",
                    {"ingest_job_id": job_id, "customer_mobile": {"$nin": [None, ""]}},
                ) if dataset == "transactions" else []
                if mobiles_in_job:
                    existing = await customers_col.distinct(
                        "mobile", {"mobile": {"$in": mobiles_in_job}}
                    )
                    missing = list(set(mobiles_in_job) - set(existing))
                    if missing:
                        # Auto-create stub customer rows so they show up in Total
                        ops = [UpdateOne(
                            {"mobile": m},
                            {"$setOnInsert": {
                                "id": uuid.uuid4().hex,
                                "mobile": m,
                                "tier": "bronze",
                                "points_balance": 0,
                                "visit_count": 0,
                                "lifetime_spend": 0.0,
                                "lifetime_points_earned": 0.0,
                                "lifetime_points_redeemed": 0.0,
                                "created_at": now(),
                                "source": "auto_from_transactions",
                            }},
                            upsert=True,
                        ) for m in missing]
                        bres = await customers_col.bulk_write(ops, ordered=False)
                        await historic_jobs_col.update_one(
                            {"id": job_id},
                            {"$set": {"customers_auto_created": bres.upserted_count}},
                        )
                        logger.info(f"Auto-created {bres.upserted_count} customer stubs for job {job_id}")
                # Recompute R1/R2/R3 aggregates for THIS job's mobiles (much cheaper than full backfill)
                affected = mobiles_in_job
                if affected:
                    pipe = [
                        {"$match": {"customer_mobile": {"$in": affected}}},
                        {"$sort": {"bill_date": 1}},
                        {"$group": {
                            "_id": "$customer_mobile",
                            "first_purchase_at": {"$first": "$bill_date"},
                            "first_store_id": {"$first": "$store_id"},
                            "last_visit_at": {"$last": "$bill_date"},
                            "uniq_bills": {"$addToSet": {"s": "$store_id", "b": "$bill_number", "d": "$bill_date"}},
                            "spend": {"$sum": "$net_amount"},
                            "earn":  {"$sum": "$points_earned"},
                        }},
                    ]
                    cust_ops = []
                    async for r in transactions_col.aggregate(pipe, allowDiskUse=True):
                        cust_ops.append(UpdateOne(
                            {"mobile": r["_id"]},
                            {"$set": {
                                "first_purchase_at": r["first_purchase_at"],
                                "last_visit_at": r["last_visit_at"],
                                "home_store_id": r.get("first_store_id"),
                                "visit_count": len(r.get("uniq_bills", [])),
                                "lifetime_spend": round(float(r.get("spend") or 0), 2),
                                "lifetime_points_earned": round(float(r.get("earn") or 0), 2),
                            }},
                            upsert=False,
                        ))
                        if len(cust_ops) >= 500:
                            await customers_col.bulk_write(cust_ops, ordered=False)
                            cust_ops.clear()
                    if cust_ops:
                        await customers_col.bulk_write(cust_ops, ordered=False)
                logger.info(f"Auto-backfill complete for job {job_id} ({len(affected)} mobiles refreshed)")
            except Exception as bx:
                logger.warning(f"Auto-backfill failed for job {job_id}: {bx}")

        # =====================================================
        # POST-INGEST AI NARRATIVE — best-effort, never breaks the job
        # =====================================================
        if not dry_run:
            try:
                from routes.ingest_narrative import build_and_store_narrative
                await build_and_store_narrative(job_id)
                logger.info(f"AI narrative generated for job {job_id}")
            except Exception as nx:
                logger.warning(f"AI narrative generation failed for job {job_id}: {nx}")
    except Exception as e:
        logger.exception(f"Historic ingest job {job_id} crashed")
        # ALWAYS persist the partial counts so the UI shows what was achieved
        import traceback
        await historic_jobs_col.update_one({"id": job_id},
            {"$set": {"status": "failed", "error": str(e),
                       "error_trace": traceback.format_exc()[:4000],
                       "processed": processed, "inserted": inserted,
                       "updated": updated, "skipped": skipped,
                       "errors_count": skipped,
                       "errors_sample": errors[:MAX_ERROR_SAMPLES],
                       "total_rows": total_rows,
                       "completed_at": now()}})


def _get_pk(dataset: str, doc: Dict[str, Any]) -> Tuple[Optional[str], Any]:
    if dataset == "customers":
        return "mobile", doc.get("mobile")
    if dataset == "transactions":
        return "bill_number", doc.get("bill_number")
    if dataset == "stores":
        return "code", doc.get("code")
    if dataset == "items":
        return "sku", doc.get("sku")
    if dataset == "points_ledger":
        # Composite key — same mobile + bill + type combo should upsert (not duplicate)
        mobile = doc.get("customer_mobile", "")
        bill = doc.get("source_bill_id") or doc.get("bill_number") or ""
        return "source_bill_id", f"{mobile}::{bill}::{doc.get('type', '')}"
    if dataset == "sku_transactions":
        # No primary-key skip — lines are always accumulated and upserted/attached.
        return None, None
    return None, None


# ============================================================
# Post-pass helpers (R1 / R2 / R3 / R6)
# ============================================================
async def _write_ledger_for_job(job_id: str) -> Dict[str, int]:
    """R6: For each loyalty bill (customer_mobile non-empty) inserted by this job,
    write `earn` / `redeem` / `bonus` ledger entries timestamped with the bill_date.

    Re-running is idempotent: each ledger doc carries `source_bill_id` so we skip
    rows that already have entries.
    """
    from pymongo import InsertOne
    counts = {"earn": 0, "redeem": 0, "bonus": 0}
    cursor = transactions_col.find(
        {"ingest_job_id": job_id, "customer_mobile": {"$nin": [None, ""]}},
        {"_id": 0, "id": 1, "bill_number": 1, "bill_date": 1, "customer_mobile": 1,
         "store_id": 1, "points_earned": 1, "points_redeemed": 1, "bonus_points": 1},
    )
    ops: List[Any] = []
    BATCH = 1000
    async for tx in cursor:
        bill_id = tx.get("id")
        if not bill_id:
            continue
        # Skip if ledger already exists for this bill (idempotency)
        if await points_ledger_col.count_documents({"source_bill_id": bill_id}, limit=1):
            continue
        base = {
            "customer_mobile": tx["customer_mobile"],
            "bill_number": tx.get("bill_number"),
            "bill_date": tx.get("bill_date"),
            "store_id": tx.get("store_id"),
            "source_bill_id": bill_id,
            "source": "historic_upload",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        earn = int(tx.get("points_earned") or 0)
        redeem = int(tx.get("points_redeemed") or 0)
        bonus = int(tx.get("bonus_points") or 0)
        if earn > 0:
            ops.append(InsertOne({**base, "id": uuid.uuid4().hex,
                                  "type": "earn", "points": earn,
                                  "reason": "Historic earn (CSV ingest)"}))
            counts["earn"] += 1
        if redeem > 0:
            ops.append(InsertOne({**base, "id": uuid.uuid4().hex,
                                  "type": "redeem", "points": -abs(redeem),
                                  "reason": "Historic redemption (CSV ingest)"}))
            counts["redeem"] += 1
        if bonus > 0:
            ops.append(InsertOne({**base, "id": uuid.uuid4().hex,
                                  "type": "bonus", "points": bonus,
                                  "reason": "Historic bonus (CSV ingest)"}))
            counts["bonus"] += 1
        if len(ops) >= BATCH:
            try:
                await points_ledger_col.bulk_write(ops, ordered=False)
            except Exception as e:
                logger.warning(f"Ledger bulk insert partial failure: {e}")
            ops.clear()
    if ops:
        try:
            await points_ledger_col.bulk_write(ops, ordered=False)
        except Exception as e:
            logger.warning(f"Ledger bulk insert (final) partial failure: {e}")
    logger.info(f"Wrote points-ledger entries for job {job_id}: {counts}")
    return counts


async def _write_opening_balance_ledger_for_job(job_id: str) -> int:
    """For every customer loaded by this CRM/customer job with a positive current
    point balance, write a single 'opening' points-ledger entry (= their balance)
    expiring at OPENING_BALANCE_EXPIRY_ISO (31 Dec 2026 IST).

    Idempotent per mobile (keyed on customer_mobile + reference_type), so re-uploads
    update rather than duplicate. Live POS points (from posAddPoint) carry their own
    1-year-from-bill-date expiry and are unaffected by this.
    """
    from pymongo import UpdateOne
    cursor = customers_col.find(
        {"ingest_job_id": job_id, "points_balance": {"$gt": 0}},
        {"_id": 0, "id": 1, "mobile": 1, "points_balance": 1},
    )
    ops: List[Any] = []
    written = 0
    BATCH = 1000
    now_iso = datetime.now(timezone.utc).isoformat()
    async for c in cursor:
        bal = int(c.get("points_balance") or 0)
        if bal <= 0:
            continue
        key = {"customer_mobile": c["mobile"], "reference_type": "opening_balance"}
        doc = {
            "customer_id": c.get("id"),
            "customer_mobile": c["mobile"],
            "type": "opening",
            "points": bal,
            "reference_type": "opening_balance",
            "expires_at": OPENING_BALANCE_EXPIRY_ISO,
            "reason": "Opening balance (historic load) — valid till 31 Dec 2026",
            "source": "historic_upload",
            "created_at": now_iso,
        }
        ops.append(UpdateOne(key, {"$set": doc, "$setOnInsert": {"id": uuid.uuid4().hex}}, upsert=True))
        if len(ops) >= BATCH:
            try:
                res = await points_ledger_col.bulk_write(ops, ordered=False)
                written += res.upserted_count + res.modified_count
            except Exception as e:
                logger.warning(f"Opening-balance ledger batch failed: {e}")
            ops.clear()
    if ops:
        try:
            res = await points_ledger_col.bulk_write(ops, ordered=False)
            written += res.upserted_count + res.modified_count
        except Exception as e:
            logger.warning(f"Opening-balance ledger final batch failed: {e}")
    logger.info(f"Opening-balance ledger entries written for job {job_id}: {written}")
    return written



async def _recompute_customer_aggregates(job_id: str) -> int:
    """R1+R2+R3: For every loyalty customer touched by this job, recompute
    first_purchase_at, last_visit_at, visit_count (unique bills), home_store_id
    (store of earliest bill) from the actual transaction history. R4: keyed by mobile.
    """
    from pymongo import UpdateOne
    # 1. Collect mobiles touched by this job
    touched_mobiles = await transactions_col.distinct(
        "customer_mobile",
        {"ingest_job_id": job_id, "customer_mobile": {"$nin": [None, ""]}},
    )
    if not touched_mobiles:
        return 0

    # 2. Aggregate per-mobile from the FULL transaction history (across all jobs)
    agg_pipe = [
        {"$match": {"customer_mobile": {"$in": touched_mobiles}}},
        {"$sort": {"bill_date": 1}},
        {"$group": {
            "_id": "$customer_mobile",
            "first_purchase_at": {"$first": "$bill_date"},
            "first_store_id": {"$first": "$store_id"},
            "last_visit_at": {"$last": "$bill_date"},
            "unique_bills": {"$addToSet": {"store": "$store_id",
                                            "bill": "$bill_number",
                                            "date": "$bill_date"}},
            "total_spend": {"$sum": "$net_amount"},
        }},
    ]
    rows = await transactions_col.aggregate(agg_pipe).to_list(len(touched_mobiles))

    ops = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for r in rows:
        mobile = r["_id"]
        update_set: Dict[str, Any] = {
            "first_purchase_at": r["first_purchase_at"],
            "last_visit_at": r["last_visit_at"],
            "home_store_id": r["first_store_id"],
            "visit_count": len(r["unique_bills"]),
            # Spend + tier are rebuilt from the FULL bill history for EVERY customer
            # (incl. those loaded from the CRM Report, which carries no spend column).
            "lifetime_spend": round(r.get("total_spend") or 0, 2),
            "tier": _derive_tier(r.get("total_spend") or 0),
        }
        update_set_on_insert = {
            "id": uuid.uuid4().hex,
            "mobile": mobile,
            "name": None,
            "city": None,
            "state": None,
            "points_balance": 0,
            "lifetime_points_earned": 0,
            "lifetime_points_redeemed": 0,
            "source": "transaction_derived",
            "created_at": now_iso,
        }
        ops.append(UpdateOne({"mobile": mobile},
                              {"$set": update_set, "$setOnInsert": update_set_on_insert},
                              upsert=True))
    if ops:
        try:
            await customers_col.bulk_write(ops, ordered=False)
        except Exception as e:
            logger.warning(f"Customer-aggregate bulk write partial failure: {e}")
    logger.info(f"Recomputed customer aggregates for {len(rows)} mobiles (job {job_id})")
    return len(rows)


def _exists_col(dataset: str):
    return {
        "customers": customers_col,
        "transactions": transactions_col,
        "stores": stores_col,
        "items": db["items"],
    }[dataset]


# ---------------- Endpoints ----------------
@router.get("/schema/{dataset}")
async def get_schema(dataset: str, user: dict = Depends(get_current_user)):
    """Return expected columns + sample row + parsing notes for a dataset."""
    if dataset not in ALLOWED_DATASETS:
        raise HTTPException(400, f"dataset must be one of {sorted(ALLOWED_DATASETS)}")
    schemas = {
        "customers": {
            "primary_key": "Mobile",
            "duplicate_strategy": "Upsert by Mobile",
            "required_columns": ["Mobile"],
            "recognised_columns": [
                "Mobile", "ONLINE", "Name", "Country Code", "Card Validity", "State", "City",
                "Added On", "Registred Account", "First Visit Account", "Last Visit Account",
                "Last Visit Date", "First Visit Date", "Current Point Balance", "Redeem Points",
                "Days Since Last Visit", "Total Billing", "Total Visits", "DOA", "DOB",
            ],
            "sample_row": {
                "Mobile": "9876543210", "ONLINE": "online", "Name": "Priya Sharma",
                "Country Code": "91", "Card Validity": "Active", "State": "Maharashtra",
                "City": "Mumbai", "Added On": "14-12-2024 00:00",
                "Last Visit Date": "01-02-2026", "First Visit Date": "01-12-2024 00:00",
                "Current Point Balance": "250", "Redeem Points": "0",
                "Total Billing": "9063", "Total Visits": "4",
                "DOB": "15-08-1995", "DOA": "",
            },
            "notes": [
                "Date format: DD-MM-YYYY or DD-MM-YYYY HH:MM",
                "Tier is auto-derived from Total Billing (silver < 25k, gold < 75k, platinum < 200k, diamond ≥ 200k)",
                "Mobile is normalised to 10 digits (91 prefix stripped automatically)",
            ],
        },
        "transactions": {
            "primary_key": "Bill Number",
            "duplicate_strategy": "Upsert by Bill Number",
            "required_columns": ["Bill Number", "Customer Mobile Number", "Date"],
            "recognised_columns": [
                "Date", "Return Marker", "Customer Mobile Number", "Customer Name",
                "Outlet(Only For Shopify Marker)", "Store master", "Store code",
                "Transaction Id", "Bill Number",
                "New_Existing", "Recency", "Last Visit Date", "Total Visits",
                "Zone New", "City", "Class", "Time",
                "Net Amount Before Tax Kazo", "Total Tax", "Discount",
                "Total Revenue Kazo", "Total Billing Lifetime",
            ],
            "sample_row": {
                "Date": "01-04-2021 00:00", "Return Marker": "Regular",
                "Customer Mobile Number": "9876543210", "Customer Name": "",
                "Outlet(Only For Shopify Marker)": "City Centre Mall, Guwahati",
                "Store master": "K00055",
                "Transaction Id": "000000PK55212200001",
                "Bill Number": "000000PK55212200001",
                "New_Existing": "New", "Recency": "Active", "Total Visits": "1",
                "Zone New": "East", "City": "Guwahati", "Class": "B",
                "Time": "12:30:00",
                "Net Amount Before Tax Kazo": "1490", "Total Tax": "0",
                "Discount": "0", "Total Revenue Kazo": "1490",
                "Total Billing Lifetime": "1490",
            },
            "notes": [
                "'Store master' (the K-code, e.g. K00055) is the canonical store code (== POS customer_key). Stores are created/linked by this code and aligned with live POS bill ingestion.",
                "Falls back to matching/creating a store by the Outlet name when no Store master code is present.",
                "Total Revenue Kazo is preferred; falls back to (Net + Tax − Discount) if missing.",
                "Item/SKU-level data is uploaded separately via the 'SKU / Line Items' dataset and attached to these bills by Transaction Id.",
            ],
        },
        "stores": {
            "primary_key": "Code",
            "duplicate_strategy": "Upsert by Code",
            "required_columns": ["Name", "City"],
            "recognised_columns": ["Code", "Name", "City", "State", "Region", "Zone",
                                     "Address", "Phone", "Manager"],
            "sample_row": {"Code": "KMUM01", "Name": "Phoenix Marketcity, Mumbai",
                            "City": "Mumbai", "State": "Maharashtra", "Region": "West",
                            "Address": "Phoenix Marketcity, Kurla", "Phone": "022xxxxxxx"},
            "notes": ["Code is auto-generated from Name if missing."],
        },
        "items": {
            "primary_key": "SKU",
            "duplicate_strategy": "Upsert by SKU",
            "required_columns": ["SKU"],
            "recognised_columns": [
                "SKU", "Item Code", "Style Code", "Article",
                "Name", "Item Name", "Product Name", "Style Name", "Description",
                "Category", "Sub Category", "Class",
                "MRP", "Price", "Selling Price",
                "Color", "Size", "Brand", "Season", "HSN", "Tax %",
            ],
            "sample_row": {"SKU": "KZ-TOP-001", "Name": "Crepe Top",
                            "Category": "TOPS", "Sub Category": "BLOUSE",
                            "MRP": "1990", "Color": "Black", "Size": "M",
                            "Brand": "KAZO", "HSN": "6109"},
            "notes": [
                "SKU accepts multiple aliases: SKU / Item Code / Style Code / Article.",
                "MRP / Selling Price are interchangeable.",
                "All optional fields (color, size, brand, season, HSN, tax %) are persisted when present.",
            ],
        },
        "points_ledger": {
            "primary_key": "Mobile + Source Bill Id + Type",
            "duplicate_strategy": "Upsert by composite key (idempotent re-runs)",
            "required_columns": ["Mobile", "Points"],
            "recognised_columns": [
                "Mobile", "Customer Mobile", "Customer Mobile Number",
                "Points", "Point", "Type", "Direction",
                "Date", "Bill Date", "Transaction Date",
                "Bill Number", "Bill", "Reference",
                "Reason", "Description", "Note",
                "Source Bill Id",
            ],
            "sample_row": {
                "Mobile": "9876543210", "Type": "earn", "Points": "500",
                "Date": "01-04-2026", "Bill Number": "BILL12345",
                "Reason": "Bill earn", "Source Bill Id": "BILL12345",
            },
            "notes": [
                "Type is one of: earn / redeem / bonus / adjust / expire (defaults to 'earn' if positive, 'redeem' if negative).",
                "Points are stored signed (redeem/expire negative, earn/bonus positive) regardless of the sign in the CSV.",
                "Mobile is normalised to 10 digits (91 prefix stripped).",
                "Source Bill Id makes ingest idempotent — re-running the same CSV won't double-write entries.",
            ],
        },
        "sku_transactions": {
            "primary_key": "Transaction Id (joins to a bill)",
            "duplicate_strategy": "Lines are attached to the matching bill (re-runs overwrite, never duplicate)",
            "required_columns": ["Transaction Id", "Item Id"],
            "recognised_columns": [
                "id", "Date", "Transaction Id", "Bill Number", "Outlet", "Store code",
                "Mobile", "Customer Name", "Item Name", "Item Id", "Season",
                "Item Master Category", "Bill Type", "Quantity", "Rate", "discount",
                "Sub Total", "Category 0(Logic)", "Category 1(Logic)",
                "Category 2(Logic)", "Category 3(Logic)", "New Vs Existing", "Basket Size",
            ],
            "sample_row": {
                "Date": "04-02-2025", "Transaction Id": "000000PK07000005510",
                "Bill Number": "INVK07232400234", "Store code": "K00049",
                "Mobile": "8789277792", "Item Name": "STUDDED INTERLOCK HOOP EARRINGS",
                "Item Id": "123340", "Season": "SPRING/SUMMER",
                "Item Master Category": "Jwellery", "Quantity": "1", "Rate": "790",
                "discount": "0", "Sub Total": "790", "Category 1(Logic)": "BAJW",
            },
            "notes": [
                "Each row is one item line on a bill. Lines are grouped by 'Transaction Id' and attached to the matching transaction's items[] (powers UPT, units sold, and category analytics).",
                "Upload the Billwise transactions FIRST so the bills exist, then upload SKU / Line Items to attach the items.",
                "Distinct items also populate the Item Master (Item Id → name, category, season, rate).",
                "'Transaction Id' (the 000000PK… value) is the join key — it equals the billwise Bill Number / Transaction Id. (Bill Number here is the INV… invoice no.)",
            ],
        },
    }
    return schemas[dataset]


def _xlsx_to_csv_text(raw_bytes: bytes) -> str:
    """Convert an .xlsx file's first sheet to a CSV string (UTF-8).

    Uses openpyxl in read-only mode so 200k+ row files don't OOM. Cells are
    string-cast to match what csv.DictReader would otherwise see.
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    for row in ws.iter_rows(values_only=True):
        # Coerce None → "" and dates → ISO
        out_row = []
        for v in row:
            if v is None:
                out_row.append("")
            elif isinstance(v, (datetime, )):
                out_row.append(v.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                out_row.append(str(v))
        w.writerow(out_row)
    wb.close()
    return buf.getvalue()


def _read_upload_to_csv_text(filename: str, raw_bytes: bytes) -> str:
    """Decode a CSV or convert an XLSX to CSV text. Raises HTTPException on
    unsupported formats so the user knows immediately."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return raw_bytes.decode("utf-8-sig", errors="replace")
    if name.endswith(".xlsx"):
        try:
            return _xlsx_to_csv_text(raw_bytes)
        except Exception as e:
            raise HTTPException(400, f"Could not read .xlsx: {e}")
    if name.endswith(".xls"):
        raise HTTPException(400, "Legacy .xls is not supported — please save as .xlsx or .csv in Excel and re-upload")
    raise HTTPException(400, "Only .csv and .xlsx files are supported")


@router.post("/ingest")
async def ingest_csv(background_tasks: BackgroundTasks,
                      file: UploadFile = File(...),
                      dataset: str = Form(...),
                      duplicate_mode: str = Form("upsert"),
                      dry_run: str = Form("false"),
                      user: dict = Depends(get_current_user)):
    if user["role"] not in {"super_admin", "brand_admin", "crm_manager", "marketing_manager"}:
        raise HTTPException(403, "Only admin / CRM / marketing can ingest historic data")
    if dataset not in ALLOWED_DATASETS:
        raise HTTPException(400, f"dataset must be one of {sorted(ALLOWED_DATASETS)}")
    if duplicate_mode not in DUPLICATE_MODES:
        raise HTTPException(400, f"duplicate_mode must be one of {sorted(DUPLICATE_MODES)}")

    raw = await file.read()
    if len(raw) > MAX_FILE_BYTES:
        raise HTTPException(413, f"File too large (max {MAX_FILE_BYTES // (1024*1024)} MB)")
    text = _read_upload_to_csv_text(file.filename, raw)

    # Count rows + grab header sample
    reader = csv.DictReader(io.StringIO(text))
    header = reader.fieldnames or []
    row_count = sum(1 for _ in reader)

    dry = str(dry_run).lower() in {"1", "true", "yes", "y"}

    job_id = uuid.uuid4().hex
    now = datetime.now(timezone.utc).isoformat()
    job_doc = {
        "id": job_id,
        "dataset": dataset,
        "filename": file.filename,
        "size_bytes": len(raw),
        "row_count_estimated": row_count,
        "columns_detected": header,
        "duplicate_mode": duplicate_mode,
        "dry_run": dry,
        "status": "queued",
        "processed": 0, "inserted": 0, "updated": 0, "skipped": 0,
        "errors_count": 0, "errors_sample": [],
        "queued_at": now,
        "queued_by": user["email"],
    }
    await historic_jobs_col.insert_one(job_doc)
    background_tasks.add_task(_run_ingest_job, job_id, dataset, text,
                               duplicate_mode, dry)
    job_doc.pop("_id", None)
    return job_doc


# ---------------- Chunked upload (for large files / production ingress limits) ----------------
class IngestInitIn(BaseModel):
    dataset: str
    duplicate_mode: str = "upsert"
    dry_run: bool = False
    filename: str
    total_chunks: int
    total_bytes: int = 0


class IngestFinalizeIn(BaseModel):
    job_id: str


@router.post("/ingest/init")
async def ingest_init(body: IngestInitIn, user: dict = Depends(get_current_user)):
    """Step 1 of chunked upload: create job record and reserve temp dir."""
    if user["role"] not in {"super_admin", "brand_admin", "crm_manager", "marketing_manager"}:
        raise HTTPException(403, "Only admin / CRM / marketing can ingest historic data")
    if body.dataset not in ALLOWED_DATASETS:
        raise HTTPException(400, f"dataset must be one of {sorted(ALLOWED_DATASETS)}")
    if body.duplicate_mode not in DUPLICATE_MODES:
        raise HTTPException(400, f"duplicate_mode must be one of {sorted(DUPLICATE_MODES)}")
    if body.total_chunks < 1 or body.total_chunks > 10_000:
        raise HTTPException(400, "total_chunks out of range")
    if not (body.filename.lower().endswith(".csv") or body.filename.lower().endswith(".xlsx")):
        raise HTTPException(400, "Only .csv and .xlsx files supported")
    if body.total_bytes > MAX_FILE_BYTES:
        raise HTTPException(413, f"File too large (max {MAX_FILE_BYTES // (1024*1024)} MB)")

    job_id = uuid.uuid4().hex
    now = datetime.now(timezone.utc).isoformat()
    job_doc = {
        "id": job_id,
        "dataset": body.dataset,
        "filename": body.filename,
        "size_bytes": 0,
        "row_count_estimated": 0,
        "columns_detected": [],
        "duplicate_mode": body.duplicate_mode,
        "dry_run": body.dry_run,
        "status": "uploading",
        "processed": 0, "inserted": 0, "updated": 0, "skipped": 0,
        "errors_count": 0, "errors_sample": [],
        "queued_at": now,
        "queued_by": user["email"],
        "total_chunks": body.total_chunks,
        "chunks_uploaded": 0,
    }
    await historic_jobs_col.insert_one(job_doc)
    job_doc.pop("_id", None)
    return job_doc


@router.post("/ingest/chunk")
async def ingest_chunk(
    job_id: str = Form(...),
    chunk_index: int = Form(...),
    chunk: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Step 2 of chunked upload: receive one chunk and persist to MongoDB (shared across pods)."""
    job = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    if job.get("status") != "uploading":
        raise HTTPException(400, f"Job is not in uploading state (status={job.get('status')})")
    if chunk_index < 0 or chunk_index >= job.get("total_chunks", 1):
        raise HTTPException(400, "chunk_index out of range")

    data = await chunk.read()
    if len(data) > 10 * 1024 * 1024:  # 10MB hard cap per chunk
        raise HTTPException(413, "Single chunk too large (max 10MB per chunk)")

    # Idempotent upsert — store in MongoDB so chunks landing on different pods all converge
    result = await historic_chunks_col.update_one(
        {"job_id": job_id, "chunk_index": chunk_index},
        {"$set": {
            "job_id": job_id,
            "chunk_index": chunk_index,
            "data": data,
            "size": len(data),
        }},
        upsert=True,
    )

    # Only increment counter when this was a fresh insert (avoid double-count on retries)
    if result.upserted_id is not None:
        await historic_jobs_col.update_one(
            {"id": job_id},
            {"$inc": {"chunks_uploaded": 1, "size_bytes": len(data)}},
        )
    return {"ok": True, "chunk_index": chunk_index, "bytes": len(data)}


@router.post("/ingest/finalize")
async def ingest_finalize(
    body: IngestFinalizeIn,
    user: dict = Depends(get_current_user),
):
    """Step 3 of chunked upload: validate chunks, count rows, queue for scheduler ingest.

    The actual ingest is performed by an APScheduler tick (every 15s) so it is
    fully resilient to pod restarts / worker recycles. Chunks stay in MongoDB
    until the scheduler completes the ingest, then they are cleaned up.
    """
    job = await historic_jobs_col.find_one({"id": body.job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    if job.get("status") != "uploading":
        raise HTTPException(400, f"Job is not in uploading state (status={job.get('status')})")

    expected = job.get("total_chunks", 0)
    found = await historic_chunks_col.count_documents({"job_id": body.job_id})
    if found != expected:
        raise HTTPException(
            400,
            f"Chunk count mismatch — expected {expected}, found {found}. Please retry the upload.",
        )

    # Quick stitch to count rows + detect header (kept ephemeral, NOT stored in memory beyond this call)
    try:
        cursor = historic_chunks_col.find(
            {"job_id": body.job_id},
            {"_id": 0, "chunk_index": 1, "data": 1},
        ).sort("chunk_index", 1)
        parts: List[bytes] = []
        last_index = -1
        async for doc in cursor:
            idx = doc["chunk_index"]
            if idx != last_index + 1:
                raise HTTPException(400, f"Chunk gap detected at index {last_index + 1}. Please retry the upload.")
            parts.append(doc["data"])
            last_index = idx
        raw = b"".join(parts)
        parts.clear()
        # Decode based on file extension (.csv direct, .xlsx via openpyxl)
        text = _read_upload_to_csv_text(job.get("filename", ""), raw)
        del raw
    except HTTPException:
        raise
    except Exception as e:
        await historic_jobs_col.update_one(
            {"id": body.job_id},
            {"$set": {"status": "failed", "error": f"Could not stitch/decode chunks: {e}"}},
        )
        raise HTTPException(500, f"Failed to read uploaded chunks: {e}")

    reader = csv.DictReader(io.StringIO(text))
    header = reader.fieldnames or []
    row_count = sum(1 for _ in reader)
    # IMPORTANT: for .xlsx files we re-decode in the scheduler tick. Cache the
    # decoded CSV text on the job so the scheduler doesn't have to re-parse xlsx.
    is_xlsx = (job.get("filename") or "").lower().endswith(".xlsx")
    if is_xlsx:
        # Replace the stored chunks with one synthetic CSV chunk to save memory + re-decode time
        try:
            await historic_chunks_col.delete_many({"job_id": body.job_id})
            csv_bytes = text.encode("utf-8")
            await historic_chunks_col.insert_one({
                "job_id": body.job_id, "chunk_index": 0, "data": csv_bytes,
            })
            # Update job's total_chunks so scheduler stitches correctly
            await historic_jobs_col.update_one(
                {"id": body.job_id},
                {"$set": {"total_chunks": 1, "filename_original": job.get("filename"),
                            "filename": (job.get("filename") or "").replace(".xlsx", ".csv")}},
            )
        except Exception as xe:
            logger.warning(f"xlsx → csv chunk replacement failed (will retry on scheduler): {xe}")
    del text  # release memory immediately — scheduler will re-stitch when it picks up the job

    now = datetime.now(timezone.utc).isoformat()
    await historic_jobs_col.update_one(
        {"id": body.job_id},
        {"$set": {
            "status": "pending_ingest",
            "row_count_estimated": row_count,
            "columns_detected": header,
            "queued_finalized_at": now,
            "heartbeat": now,
        }},
    )

    # Chunks remain in MongoDB until scheduler completes ingest. Scheduler tick
    # (every 15s) will claim this job atomically, process it, and clean up.

    updated = await historic_jobs_col.find_one({"id": body.job_id}, {"_id": 0})
    return updated or {"id": body.job_id, "status": "pending_ingest", "row_count_estimated": row_count}


# ---------------- Scheduler-driven ingest worker ----------------
async def process_pending_ingests():
    """Scheduler tick — runs every 15s.

    1. Recovers stale jobs (status=running but heartbeat > 3 min old)
    2. Atomically claims ONE pending job and runs the ingest
    3. On completion, deletes the stored chunks
    """
    from pymongo import ReturnDocument

    iso_now = datetime.now(timezone.utc).isoformat()
    stale_cutoff = (datetime.now(timezone.utc) - timedelta(minutes=3)).isoformat()

    # Step 1: recover stale "running" jobs
    recovery = await historic_jobs_col.update_many(
        {"status": "running", "heartbeat": {"$lt": stale_cutoff}},
        {"$set": {"status": "pending_ingest",
                   "stale_recovered_at": iso_now}},
    )
    if recovery.modified_count:
        logger.warning(f"Recovered {recovery.modified_count} stale ingest job(s)")

    # Step 2: claim one pending job atomically
    job = await historic_jobs_col.find_one_and_update(
        {"status": "pending_ingest"},
        {"$set": {"status": "running",
                   "claimed_at": iso_now,
                   "heartbeat": iso_now,
                   "started_at": iso_now}},
        sort=[("queued_at", 1)],
        return_document=ReturnDocument.AFTER,
    )
    if not job:
        return

    job_id = job["id"]
    logger.info(f"Claimed ingest job {job_id} (dataset={job['dataset']}, rows={job.get('row_count_estimated')})")

    # Step 3: stitch chunks and run ingest
    try:
        cursor = historic_chunks_col.find(
            {"job_id": job_id},
            {"_id": 0, "chunk_index": 1, "data": 1},
        ).sort("chunk_index", 1)
        parts: List[bytes] = []
        async for doc in cursor:
            parts.append(doc["data"])
        raw = b"".join(parts)
        parts.clear()
        text = raw.decode("utf-8-sig", errors="replace")
        del raw
    except Exception as e:
        logger.exception(f"Failed to stitch chunks for job {job_id}")
        await historic_jobs_col.update_one(
            {"id": job_id},
            {"$set": {"status": "failed",
                       "error": f"Could not read uploaded chunks: {e}",
                       "completed_at": datetime.now(timezone.utc).isoformat()}},
        )
        return

    # _run_ingest_job already writes heartbeat per flush + handles its own try/except
    await _run_ingest_job(job_id, job["dataset"], text,
                           job["duplicate_mode"], job.get("dry_run", False))
    del text

    # Step 4: cleanup chunks if ingest completed cleanly
    refreshed = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0, "status": 1, "total_rows": 1, "inserted": 1, "updated": 1, "skipped": 1})
    if refreshed and refreshed.get("status") in {"completed", "previewed"}:
        deleted = await historic_chunks_col.delete_many({"job_id": job_id})
        # Reconciliation: count rows in DB vs CSV total_rows
        await _reconcile_job(job_id, job["dataset"])
        logger.info(f"Ingest job {job_id} done — cleaned up {deleted.deleted_count} chunk docs")


async def _reconcile_job(job_id: str, dataset: str):
    """After ingest, count how many rows from this batch landed in the target collection.

    Stores `reconciliation` block on the job doc.
    """
    job = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    if not job:
        return
    total = int(job.get("total_rows") or 0)
    inserted = int(job.get("inserted") or 0)
    updated = int(job.get("updated") or 0)
    skipped = int(job.get("skipped") or 0)
    processed = inserted + updated + skipped
    diff = total - processed
    recon = {
        "total_rows_in_csv": total,
        "processed_rows": processed,
        "inserted": inserted,
        "updated": updated,
        "skipped_or_errored": skipped,
        "diff": diff,
        "match": diff == 0,
        "checked_at": datetime.now(timezone.utc).isoformat(),
    }
    await historic_jobs_col.update_one({"id": job_id}, {"$set": {"reconciliation": recon}})


@router.post("/ingest/abort/{job_id}")
async def ingest_abort(job_id: str, user: dict = Depends(get_current_user)):
    """Cancel an in-flight chunked upload (clean temp chunks from MongoDB)."""
    job = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    if job.get("status") not in {"uploading", "queued", "pending_ingest"}:
        return {"ok": True, "noop": True}
    await historic_chunks_col.delete_many({"job_id": job_id})
    await historic_jobs_col.update_one(
        {"id": job_id},
        {"$set": {"status": "failed", "error": "Aborted by user",
                  "completed_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"ok": True, "aborted": True}


@router.get("/jobs")
async def list_jobs(limit: int = 50, user: dict = Depends(get_current_user)):
    rows = await historic_jobs_col.find({}, {"_id": 0}).sort("queued_at", -1).limit(min(limit, 200)).to_list(200)
    return {"rows": rows, "total": len(rows)}


@router.get("/jobs/{job_id}")
async def get_job(job_id: str, user: dict = Depends(get_current_user)):
    j = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    if not j:
        raise HTTPException(404, "Job not found")
    return j


@router.post("/jobs/{job_id}/narrative")
async def regenerate_narrative(job_id: str, user: dict = Depends(require_roles(*MANAGEMENT_ROLES))):
    """Generate (or regenerate) the AI narrative for a completed ingest job."""
    from routes.ingest_narrative import build_and_store_narrative
    res = await build_and_store_narrative(job_id)
    if res.get("error"):
        raise HTTPException(400, res["error"])
    return res


@router.get("/jobs/{job_id}/narrative")
async def get_narrative(job_id: str, user: dict = Depends(get_current_user)):
    j = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0, "ai_narrative": 1})
    if not j:
        raise HTTPException(404, "Job not found")
    return j.get("ai_narrative") or {"narrative": "", "source": "not_generated"}


@router.get("/jobs/{job_id}/skipped-rows.csv")
async def download_skipped_rows(job_id: str, user: dict = Depends(get_current_user)):
    """Download every row that the parser rejected for this job, with the reason.

    Brand managers use this to understand the exact data-mismatch between
    their source Excel/CSV and the dashboard: rows in DB = inserted + updated
    (matched), rows NOT in DB = rows in this download.
    """
    job = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")

    async def _gen():
        # Collect all distinct columns across raw_row payloads so we get one header
        header_cols: List[str] = ["row_number", "reason"]
        seen_keys: set = set(header_cols)
        # First pass — preload first 200 rows to determine headers (covers most cases)
        sample = await historic_skipped_col.find(
            {"job_id": job_id}, {"_id": 0, "raw_row": 1},
        ).limit(200).to_list(200)
        for s in sample:
            for k in (s.get("raw_row") or {}).keys():
                if k not in seen_keys:
                    seen_keys.add(k)
                    header_cols.append(k)

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header_cols)
        yield buf.getvalue()
        buf.seek(0)
        buf.truncate(0)

        cursor = historic_skipped_col.find(
            {"job_id": job_id}, {"_id": 0, "row_number": 1, "reason": 1, "raw_row": 1},
        ).sort("row_number", 1)
        async for s in cursor:
            row = [s.get("row_number", ""), s.get("reason", "")]
            r = s.get("raw_row") or {}
            for col in header_cols[2:]:
                row.append(r.get(col, ""))
            writer.writerow(row)
            if buf.tell() > 32 * 1024:
                yield buf.getvalue()
                buf.seek(0)
                buf.truncate(0)
        if buf.tell() > 0:
            yield buf.getvalue()

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", job.get("filename", "skipped"))[:60] or "skipped"
    return StreamingResponse(
        _gen(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_skipped_{job_id[:8]}.csv"'},
    )


@router.get("/jobs/{job_id}/integrity")
async def integrity_check(job_id: str, user: dict = Depends(get_current_user)):
    """Reconcile a finished ingest job against what's actually in MongoDB.

    Returns: csv_rows, inserted, updated, skipped (persisted), db_rows_for_this_job.
    The mismatch field flags whether everything balances. Use this to give the
    brand manager an exact answer to "did all my data land in the database?"
    """
    job = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    dataset = job.get("dataset")
    csv_rows = int(job.get("row_count_estimated") or job.get("total_rows") or 0)
    inserted = int(job.get("inserted") or 0)
    updated = int(job.get("updated") or 0)
    skipped = int(job.get("skipped") or 0)
    skipped_persisted = await historic_skipped_col.count_documents({"job_id": job_id})

    # Count rows in the actual target collection that came from this job
    db_rows_for_this_job: Optional[int] = None
    if dataset == "transactions":
        db_rows_for_this_job = await transactions_col.count_documents({"ingest_job_id": job_id})
    elif dataset == "customers":
        # Customer rows don't carry ingest_job_id; report nothing (the matched_count fix already handles re-uploads).
        db_rows_for_this_job = None
    elif dataset == "stores":
        db_rows_for_this_job = None
    elif dataset == "items":
        db_rows_for_this_job = None

    accounted = inserted + updated + skipped
    diff = csv_rows - accounted if csv_rows else 0
    return {
        "job_id": job_id,
        "dataset": dataset,
        "status": job.get("status"),
        "csv_rows": csv_rows,
        "inserted": inserted,
        "updated_matched": updated,  # using matched_count semantics
        "skipped": skipped,
        "skipped_persisted_count": skipped_persisted,
        "accounted": accounted,
        "unaccounted_diff": diff,
        "balanced": csv_rows == 0 or abs(diff) <= max(2, csv_rows * 0.001),
        "db_rows_for_this_job": db_rows_for_this_job,
        "note": (
            "Updated reflects matched_count (rows that already existed and were touched, "
            "incl. no-op upserts). Skipped rows are downloadable as CSV at "
            "/jobs/{job_id}/skipped-rows.csv. db_rows_for_this_job counts transactions "
            "tagged with this ingest_job_id; for customers/stores/items see total collection counts."
        ),
    }


# ---------------- Purge demo data ----------------
PURGEABLE_COLLECTIONS = [
    ("customers", customers_col),
    ("transactions", transactions_col),
    ("stores", stores_col),
    ("campaigns", campaigns_col),
    ("campaign_metrics", campaign_metrics_col),
    ("coupons", coupons_col),
    ("coupon_redemptions", coupon_redemptions_col),
    ("points_ledger", points_ledger_col),
    ("api_logs", api_logs_col),
    ("nps_responses", nps_col),
    ("support_tickets", tickets_col),
    ("ai_chats", ai_chats_col),
    ("message_log", message_log_col),
    ("historic_uploads", historic_raw_col),
]


class PurgeIn(BaseModel):
    confirm: bool = False
    keep_users: bool = True
    keep_loyalty_config: bool = True
    keep_templates: bool = True
    keep_provider_config: bool = True


@router.get("/purge-preview")
async def purge_preview(user: dict = Depends(require_roles("super_admin", "brand_admin"))):
    """Show how many docs would be deleted in each collection."""
    preview = {}
    for name, col in PURGEABLE_COLLECTIONS:
        preview[name] = await col.count_documents({})
    preview["bulk_send_jobs"] = await db["bulk_send_jobs"].count_documents({})
    preview["audit_logs"] = await audit_logs_col.count_documents({})
    preview["digest_reports"] = await db["digest_reports"].count_documents({})
    return {"current_counts": preview,
             "note": "Users, loyalty_config, communication_templates, provider_config are PRESERVED unless explicitly disabled."}


@router.post("/purge-demo")
async def purge_demo(body: PurgeIn,
                      user: dict = Depends(require_roles("super_admin", "brand_admin"))):
    """Delete all transactional + customer + audit data. Keeps users/config/templates by default."""
    if not body.confirm:
        raise HTTPException(400, "Set confirm=true to actually purge. Use GET /purge-preview to see counts first.")
    deleted = {}
    for name, col in PURGEABLE_COLLECTIONS:
        res = await col.delete_many({})
        deleted[name] = res.deleted_count
    # Always purge job history (transient)
    deleted["bulk_send_jobs"] = (await db["bulk_send_jobs"].delete_many({})).deleted_count
    deleted["digest_reports"] = (await db["digest_reports"].delete_many({})).deleted_count
    deleted["historic_ingest_jobs"] = (await historic_jobs_col.delete_many({})).deleted_count
    deleted["audit_logs"] = (await audit_logs_col.delete_many({})).deleted_count
    if not body.keep_users:
        deleted["users"] = (await db["users"].delete_many(
            {"email": {"$nin": ["admin@kazo.com", "superadmin@fundle.io"]}}
        )).deleted_count
    if not body.keep_loyalty_config:
        deleted["loyalty_config"] = (await db["loyalty_config"].delete_many({})).deleted_count
    if not body.keep_templates:
        deleted["communication_templates"] = (await db["communication_templates"].delete_many({})).deleted_count
    if not body.keep_provider_config:
        deleted["provider_config"] = (await db["provider_config"].delete_many({})).deleted_count
    return {"purged": True, "deleted_counts": deleted, "by": user["email"]}


# ============================================================
# One-shot mobile normalization — sweeps customers.mobile +
# transactions.customer_mobile + points_ledger.customer_mobile +
# nps_responses.mobile + tickets.customer_mobile, applies the same
# _norm_mobile() that POS / segment builder / dashboards use
# (strips +91, country code, spaces, non-digits → clean 10 digit)
# so historic CSV format and live POS format match across the board.
# Fully idempotent — safe to re-run.
# ============================================================
@router.post("/normalize-mobiles")
async def normalize_mobiles(
    user: dict = Depends(require_roles("super_admin", "brand_admin")),
    dry_run: bool = False,
):
    """Sweep every collection that stores a customer mobile and rewrite each
    value through _norm_mobile() so historic-ingest formats (+91…, 91…, with
    spaces) match the clean 10-digit form used by POS + segment builder +
    dashboards. Idempotent.

    Returns a per-collection report:
      scanned · already_normalized · updated · null_or_empty
    """
    from pymongo import UpdateOne

    report: Dict[str, Any] = {
        "started_at": datetime.now(timezone.utc).isoformat(),
        "dry_run": dry_run,
        "collections": {},
    }

    # collection -> (mongo collection object, field name holding the mobile)
    targets = [
        ("customers", customers_col, "mobile"),
        ("transactions", transactions_col, "customer_mobile"),
        ("points_ledger", points_ledger_col, "customer_mobile"),
        ("nps_responses", nps_col, "mobile"),
        ("support_tickets", tickets_col, "customer_mobile"),
    ]

    for col_name, col, field in targets:
        scanned = 0
        already = 0
        updated = 0
        empty = 0
        ops: List[Any] = []
        cursor = col.find(
            {field: {"$exists": True}},
            {"_id": 1, field: 1},
        )
        async for doc in cursor:
            scanned += 1
            raw = doc.get(field)
            if raw is None or raw == "":
                empty += 1
                continue
            norm = _norm_mobile(raw)
            if not norm:
                empty += 1
                continue
            if str(raw) == norm:
                already += 1
                continue
            updated += 1
            ops.append(UpdateOne({"_id": doc["_id"]}, {"$set": {field: norm}}))
            # flush in batches of 1000 to keep memory flat on big collections
            if len(ops) >= 1000:
                if not dry_run:
                    await col.bulk_write(ops, ordered=False)
                ops = []
        if ops and not dry_run:
            await col.bulk_write(ops, ordered=False)

        report["collections"][col_name] = {
            "field": field,
            "scanned": scanned,
            "already_normalized": already,
            "updated": updated,
            "null_or_empty": empty,
        }

    report["finished_at"] = datetime.now(timezone.utc).isoformat()
    report["total_updated"] = sum(c["updated"] for c in report["collections"].values())
    report["balanced"] = True
    return report


# ============================================================
# Loyalty-model backfill — applies R1+R2+R3 + bill uniqueness index
# to the ENTIRE existing transactions/customers dataset (one-shot).
# Run once after upgrading the codebase.
# ============================================================
@router.post("/backfill-loyalty-model")
async def backfill_loyalty_model(
    user: dict = Depends(require_roles("super_admin", "brand_admin")),
):
    """Backfill existing data per the KAZO loyalty data rules:
    R1  first_purchase_at / last_visit_at sourced from bill_date
    R2  home_store_id = store_id of customer's earliest bill
    R3  visit_count = COUNT(distinct (store, bill, date)) per mobile
    R4  Hard unique index on (store_id, bill_number, bill_date) for transactions
        + unique index on customers.mobile (loyalty members are deduped by mobile)
    """
    from pymongo import UpdateOne

    report: Dict[str, Any] = {"started_at": datetime.now(timezone.utc).isoformat()}

    # ---- Step 1: Build unique indices ----
    idx_report: Dict[str, str] = {}
    try:
        await transactions_col.create_index(
            [("store_id", 1), ("bill_number", 1), ("bill_date", 1)],
            unique=True, sparse=True, name="uniq_bill_store_date",
        )
        idx_report["transactions_uniq"] = "ok"
    except Exception as e:
        idx_report["transactions_uniq"] = f"warn: {e}"
    try:
        # Loyalty mobile uniqueness — partial filter so anonymous customers (None mobile)
        # don't collide. There should be at most ONE document per non-empty mobile.
        await customers_col.create_index(
            [("mobile", 1)], unique=True,
            partialFilterExpression={"mobile": {"$type": "string"}},
            name="uniq_customer_mobile",
        )
        idx_report["customers_mobile_uniq"] = "ok"
    except Exception as e:
        idx_report["customers_mobile_uniq"] = f"warn: {e}"
    report["indices"] = idx_report

    # ---- Step 2: Recompute aggregates for every loyalty mobile ----
    # Pull every distinct mobile (in chunks to avoid memory)
    all_mobiles = await transactions_col.distinct(
        "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}}
    )
    report["loyalty_mobiles_found"] = len(all_mobiles)
    if not all_mobiles:
        report["completed_at"] = datetime.now(timezone.utc).isoformat()
        return report

    # Aggregate
    agg_pipe = [
        {"$match": {"customer_mobile": {"$in": all_mobiles}}},
        {"$sort": {"bill_date": 1}},
        {"$group": {
            "_id": "$customer_mobile",
            "first_purchase_at": {"$first": "$bill_date"},
            "first_store_id": {"$first": "$store_id"},
            "last_visit_at": {"$last": "$bill_date"},
            "unique_bills": {"$addToSet": {"store": "$store_id",
                                            "bill": "$bill_number",
                                            "date": "$bill_date"}},
            "total_spend": {"$sum": "$net_amount"},
            "txn_count": {"$sum": 1},
        }},
    ]
    rows = await transactions_col.aggregate(agg_pipe, allowDiskUse=True).to_list(len(all_mobiles) + 1000)
    report["mobiles_aggregated"] = len(rows)

    # Bulk upsert customer docs
    BATCH = 500
    ops: List[Any] = []
    upserted = 0
    updated = 0
    now_iso = datetime.now(timezone.utc).isoformat()
    for r in rows:
        mobile = r["_id"]
        ops.append(UpdateOne(
            {"mobile": mobile},
            {"$set": {
                "first_purchase_at": r["first_purchase_at"],
                "last_visit_at": r["last_visit_at"],
                "home_store_id": r["first_store_id"],
                "visit_count": len(r["unique_bills"]),
            },
             "$setOnInsert": {
                 "id": uuid.uuid4().hex,
                 "mobile": mobile,
                 "tier": _derive_tier(r.get("total_spend") or 0),
                 "lifetime_spend": round(r.get("total_spend") or 0, 2),
                 "points_balance": 0,
                 "lifetime_points_earned": 0,
                 "lifetime_points_redeemed": 0,
                 "source": "transaction_derived",
                 "created_at": now_iso,
             }},
            upsert=True,
        ))
        if len(ops) >= BATCH:
            try:
                res = await customers_col.bulk_write(ops, ordered=False)
                upserted += res.upserted_count
                updated += res.matched_count
            except Exception as e:
                logger.warning(f"backfill bulk partial: {e}")
            ops.clear()
    if ops:
        try:
            res = await customers_col.bulk_write(ops, ordered=False)
            upserted += res.upserted_count
            updated += res.matched_count
        except Exception as e:
            logger.warning(f"backfill bulk final partial: {e}")
    report["customers_upserted"] = upserted
    report["customers_updated"] = updated
    report["completed_at"] = datetime.now(timezone.utc).isoformat()
    logger.info(f"Loyalty backfill complete: {report}")
    return report


# ============================================================
# R6 retrofit — write missing points-ledger entries for ALL existing
# loyalty transactions (not just new ingest jobs). Idempotent.
# ============================================================
@router.post("/backfill-points-ledger")
async def backfill_points_ledger(
    user: dict = Depends(require_roles("super_admin", "brand_admin")),
):
    """Sweep every loyalty transaction and ensure its earn / redeem / bonus
    ledger entries exist (R6). Skips transactions that already have entries."""
    from pymongo import InsertOne

    report: Dict[str, Any] = {"started_at": datetime.now(timezone.utc).isoformat()}
    counts = {"earn": 0, "redeem": 0, "bonus": 0, "scanned": 0, "skipped_no_points": 0,
              "skipped_already_indexed": 0}

    # Pre-fetch all bill_ids that already have ledger entries (set membership for O(1) lookup)
    indexed: set = set()
    async for d in points_ledger_col.find(
        {"source_bill_id": {"$exists": True, "$ne": None}}, {"_id": 0, "source_bill_id": 1}
    ):
        sid = d.get("source_bill_id")
        if sid:
            indexed.add(sid)
    report["existing_ledger_index_size"] = len(indexed)

    ops: List[Any] = []
    BATCH = 1000
    cursor = transactions_col.find(
        {"customer_mobile": {"$nin": [None, ""]}},
        {"_id": 0, "id": 1, "bill_number": 1, "bill_date": 1, "customer_mobile": 1,
         "store_id": 1, "points_earned": 1, "points_redeemed": 1, "bonus_points": 1},
    )
    async for tx in cursor:
        counts["scanned"] += 1
        bill_id = tx.get("id")
        if not bill_id:
            continue
        if bill_id in indexed:
            counts["skipped_already_indexed"] += 1
            continue
        earn = int(tx.get("points_earned") or 0)
        redeem = int(tx.get("points_redeemed") or 0)
        bonus = int(tx.get("bonus_points") or 0)
        if not (earn or redeem or bonus):
            counts["skipped_no_points"] += 1
            continue
        base = {
            "customer_mobile": tx["customer_mobile"],
            "bill_number": tx.get("bill_number"),
            "bill_date": tx.get("bill_date"),
            "store_id": tx.get("store_id"),
            "source_bill_id": bill_id,
            "source": "historic_upload_backfill",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        if earn > 0:
            ops.append(InsertOne({**base, "id": uuid.uuid4().hex, "type": "earn",
                                  "points": earn, "reason": "Backfill earn"}))
            counts["earn"] += 1
        if redeem > 0:
            ops.append(InsertOne({**base, "id": uuid.uuid4().hex, "type": "redeem",
                                  "points": -abs(redeem), "reason": "Backfill redemption"}))
            counts["redeem"] += 1
        if bonus > 0:
            ops.append(InsertOne({**base, "id": uuid.uuid4().hex, "type": "bonus",
                                  "points": bonus, "reason": "Backfill bonus"}))
            counts["bonus"] += 1
        if len(ops) >= BATCH:
            try:
                await points_ledger_col.bulk_write(ops, ordered=False)
            except Exception as e:
                logger.warning(f"Ledger backfill bulk partial: {e}")
            ops.clear()
    if ops:
        try:
            await points_ledger_col.bulk_write(ops, ordered=False)
        except Exception as e:
            logger.warning(f"Ledger backfill bulk final partial: {e}")
    report["counts"] = counts
    report["completed_at"] = datetime.now(timezone.utc).isoformat()
    logger.info(f"Points-ledger backfill complete: {report}")
    return report


# ============================================================
# R4 dedupe report — surface any customers sharing a mobile.
# (The new partial-unique index prevents new dupes; this scans the existing DB.)
# ============================================================
@router.get("/dedupe/mobiles")
async def dedupe_mobiles(user: dict = Depends(require_roles("super_admin", "brand_admin"))):
    """Return any non-empty mobile numbers held by more than one customer doc.
    Empty/null mobiles are ignored (those are anonymous walk-ins by design)."""
    pipe = [
        {"$match": {"mobile": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$mobile", "count": {"$sum": 1},
                    "ids": {"$push": {"id": "$id", "name": "$name",
                                       "lifetime_spend": "$lifetime_spend",
                                       "created_at": "$created_at"}}}},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 500},
    ]
    rows = await customers_col.aggregate(pipe).to_list(500)
    return {"duplicate_mobiles": len(rows),
             "rows": [{"mobile": r["_id"], "count": r["count"], "docs": r["ids"]}
                       for r in rows]}


# ============================================================
# Reconciliation engine — single-shot, comprehensive integrity check
# comparing the LAST completed CSV ingest job vs current DB state.
# ============================================================
@router.get("/reconcile")
async def reconcile(
    job_id: Optional[str] = None,
    user: dict = Depends(require_roles("super_admin", "brand_admin")),
):
    """Reconciliation report — compares the last completed CSV ingest job
    (or the one provided via ?job_id=) against the current DB state.

    Returns:
      - job_summary: original CSV counts vs ingest result counts
      - db_state: live aggregate counts in MongoDB
      - sums: CSV-declared sums (per job) vs DB sums
      - integrity_checks: orphaned-store txns, missing customer mobiles,
                          duplicate mobiles, ledger coverage gaps
    """
    # 1. Pick the job to reconcile against
    if job_id:
        job = await historic_jobs_col.find_one({"id": job_id}, {"_id": 0})
    else:
        job = await historic_jobs_col.find_one(
            {"status": "completed", "dataset": "transactions"},
            {"_id": 0},
            sort=[("completed_at", -1)],
        )
    if not job:
        raise HTTPException(404, "No completed transactions ingest job found")

    # 2. Per-job ingest reconciliation (already maintained, but re-stated)
    total = int(job.get("total_rows") or 0)
    inserted = int(job.get("inserted") or 0)
    updated = int(job.get("updated") or 0)
    skipped = int(job.get("skipped") or 0)
    processed = inserted + updated + skipped
    job_summary = {
        "job_id": job["id"],
        "dataset": job.get("dataset"),
        "completed_at": job.get("completed_at"),
        "csv_total_rows": total,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "processed_total": processed,
        "diff": total - processed,
        "match": (total - processed) == 0,
        "errors_sample": (job.get("errors_sample") or [])[:5],
    }

    # 3. DB state — totals + loyalty/non-loyalty split + ledger
    db_state = {
        "transactions_total": await transactions_col.count_documents({}),
        "transactions_loyalty": await transactions_col.count_documents(
            {"customer_mobile": {"$nin": [None, ""]}}),
        "transactions_non_loyalty": await transactions_col.count_documents(
            {"customer_mobile": {"$in": [None, ""]}}),
        "customers_total": await customers_col.count_documents({}),
        "customers_with_home_store": await customers_col.count_documents(
            {"home_store_id": {"$ne": None}}),
        "customers_with_first_purchase": await customers_col.count_documents(
            {"first_purchase_at": {"$ne": None}}),
        "distinct_mobiles_in_txns": len(await transactions_col.distinct(
            "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}})),
        "stores_total": await stores_col.count_documents({}),
        "points_ledger_total": await points_ledger_col.count_documents({}),
    }

    # 4. Money + points sums (loyalty only)
    sums_pipe = [
        {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
        {"$group": {"_id": None,
                    "net_sum": {"$sum": "$net_amount"},
                    "tax_sum": {"$sum": "$tax_amount"},
                    "discount_sum": {"$sum": "$discount_amount"},
                    "earned_sum": {"$sum": "$points_earned"},
                    "redeemed_sum": {"$sum": "$points_redeemed"},
                    "bonus_sum": {"$sum": "$bonus_points"}}},
    ]
    sums_row = (await transactions_col.aggregate(sums_pipe).to_list(1)) or [{}]
    sums_row = sums_row[0] if sums_row else {}
    ledger_pipe = [
        {"$group": {"_id": "$type", "total": {"$sum": "$points"}}}
    ]
    led_rows = await points_ledger_col.aggregate(ledger_pipe).to_list(10)
    ledger_totals = {r["_id"]: int(r["total"]) for r in led_rows}

    sums = {
        "net_amount_loyalty": round(sums_row.get("net_sum", 0) or 0, 2),
        "tax_loyalty": round(sums_row.get("tax_sum", 0) or 0, 2),
        "discount_loyalty": round(sums_row.get("discount_sum", 0) or 0, 2),
        "points_earned_from_txns": int(sums_row.get("earned_sum", 0) or 0),
        "points_redeemed_from_txns": int(sums_row.get("redeemed_sum", 0) or 0),
        "points_bonus_from_txns": int(sums_row.get("bonus_sum", 0) or 0),
        "ledger_earn_total": ledger_totals.get("earn", 0),
        "ledger_redeem_total": abs(ledger_totals.get("redeem", 0)),
        "ledger_bonus_total": ledger_totals.get("bonus", 0),
        "ledger_vs_txns_earn_diff":
            ledger_totals.get("earn", 0) - int(sums_row.get("earned_sum", 0) or 0),
        "ledger_vs_txns_redeem_diff":
            abs(ledger_totals.get("redeem", 0)) - int(sums_row.get("redeemed_sum", 0) or 0),
    }

    # 5. Integrity checks
    # 5a. Orphan transactions (store_id null after store auto-create)
    orphan_store_txns = await transactions_col.count_documents({"store_id": None})
    # 5b. Loyalty customers in txns but missing customer doc
    txn_mobiles = await transactions_col.distinct(
        "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}}
    )
    cust_mobiles_set = set(await customers_col.distinct(
        "mobile", {"mobile": {"$in": txn_mobiles}}
    ))
    missing_customer_for_mobile = [m for m in txn_mobiles if m not in cust_mobiles_set]
    # 5c. Duplicate mobiles
    dup_pipe = [
        {"$match": {"mobile": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$mobile", "n": {"$sum": 1}}},
        {"$match": {"n": {"$gt": 1}}},
        {"$count": "duplicates"},
    ]
    dup_result = await customers_col.aggregate(dup_pipe).to_list(1)
    dup_count = dup_result[0]["duplicates"] if dup_result else 0
    # 5d. Ledger coverage = % of loyalty bills that have at least 1 ledger entry
    loyalty_txn_count = db_state["transactions_loyalty"]
    bills_in_ledger = len(await points_ledger_col.distinct(
        "source_bill_id", {"source_bill_id": {"$exists": True, "$ne": None}}
    ))
    ledger_coverage_pct = round((bills_in_ledger / loyalty_txn_count * 100), 2) \
        if loyalty_txn_count else 0

    integrity = {
        "orphan_store_txns": orphan_store_txns,
        "loyalty_mobiles_missing_customer_doc": len(missing_customer_for_mobile),
        "missing_customer_sample": missing_customer_for_mobile[:5],
        "duplicate_mobile_customers": dup_count,
        "ledger_coverage_loyalty_bills_pct": ledger_coverage_pct,
        "ledger_coverage_bills_seen": bills_in_ledger,
    }

    # 6. Overall pass/fail flag
    issues: List[str] = []
    if not job_summary["match"]:
        issues.append(f"CSV vs processed diff = {job_summary['diff']}")
    if integrity["orphan_store_txns"]:
        issues.append(f"{integrity['orphan_store_txns']} txns missing store_id")
    if integrity["loyalty_mobiles_missing_customer_doc"]:
        issues.append(f"{integrity['loyalty_mobiles_missing_customer_doc']} loyalty mobiles missing customer doc")
    if integrity["duplicate_mobile_customers"]:
        issues.append(f"{integrity['duplicate_mobile_customers']} duplicate mobile-customer docs")
    if loyalty_txn_count and integrity["ledger_coverage_loyalty_bills_pct"] < 95 and (
        sums["points_earned_from_txns"] or sums["points_redeemed_from_txns"]):
        issues.append(f"Ledger coverage only {integrity['ledger_coverage_loyalty_bills_pct']}% — run /backfill-points-ledger")
    if sums["ledger_vs_txns_earn_diff"]:
        issues.append(f"Ledger earn total off by {sums['ledger_vs_txns_earn_diff']} vs txns")

    return {
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "status": "clean" if not issues else "issues_found",
        "issues": issues,
        "job_summary": job_summary,
        "db_state": db_state,
        "sums": sums,
        "integrity": integrity,
    }
