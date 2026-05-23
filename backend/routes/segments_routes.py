"""Segment Builder — exhaustive customer-cohort filters for the Campaign Manager.

Implements KAZO's locked loyalty data rules (R1-R6).

Endpoints (mounted under /api/segments):
  GET    /filter-schema      → full filter taxonomy (frontend uses this to render rows)
  POST   /preview            → live count + reach breakdown for a filter tree
  POST   /facets             → distinct values for type-ahead dropdowns
  POST   /                   → save a named segment
  GET    /                   → list saved segments
  GET    /{id}               → fetch one
  PUT    /{id}               → update (creator only)
  DELETE /{id}               → delete (creator only)
  POST   /{id}/refresh       → recompute its cached counts

Filter tree shape (max 2 levels of nesting):
{
  "op": "AND" | "OR",
  "rules": [
    { "field": "tier", "operator": "in", "value": ["gold","platinum"] },
    {
      "op": "OR",
      "rules": [
        { "field": "lifetime_spend", "operator": "gte", "value": 25000 },
        { "field": "visit_count",   "operator": "gte", "value": 5 }
      ]
    }
  ]
}
"""
from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from auth import get_current_user, require_roles
from database import (
    customers_col, transactions_col, stores_col, coupon_redemptions_col,
    nps_col, tickets_col, message_log_col, db,
)

router = APIRouter(prefix="/segments", tags=["segments"])

segments_col = db["segments"]


# ============================================================
# Filter taxonomy — single source of truth.
# Frontend hits /filter-schema; preview/save endpoints validate against this.
# ============================================================
def _now() -> datetime:
    return datetime.now(timezone.utc)


def _iso(d: datetime) -> str:
    return d.isoformat()


# Operator metadata — used by both frontend (to render UI) and backend (to compile)
OP_NUMERIC = ["eq", "neq", "gte", "lte", "between"]
OP_MULTI = ["in", "not_in"]
OP_BOOL = ["eq"]
OP_DATE = ["gte", "lte", "between"]


FILTER_SCHEMA: Dict[str, Dict[str, Any]] = {
    # ---------- 📍 Geography ----------
    "geography": {
        "label": "Geography",
        "icon": "MapPin",
        "fields": [
            {"key": "city", "label": "Customer city", "type": "multi_async",
             "operators": OP_MULTI, "facet": "customers.city"},
            {"key": "state", "label": "Customer state", "type": "multi_async",
             "operators": OP_MULTI, "facet": "customers.state"},
            {"key": "country_code", "label": "Country code", "type": "multi_async",
             "operators": OP_MULTI, "facet": "customers.country_code"},
            {"key": "home_store_id", "label": "Home store (first-bill store · R2)",
             "type": "multi_async", "operators": OP_MULTI, "facet": "stores"},
            {"key": "home_store_region", "label": "Home store region",
             "type": "multi_async", "operators": OP_MULTI, "facet": "stores.region"},
            {"key": "home_store_city", "label": "Home store city",
             "type": "multi_async", "operators": OP_MULTI, "facet": "stores.city"},
        ],
    },
    # ---------- 👤 Identity & Profile ----------
    "identity": {
        "label": "Identity & Profile",
        "icon": "UserRound",
        "fields": [
            {"key": "gender", "label": "Gender", "type": "multi",
             "operators": OP_MULTI, "options": ["M", "F", "Other"]},
            {"key": "age_band", "label": "Age band (from DOB)", "type": "multi",
             "operators": OP_MULTI,
             "options": ["18-24", "25-34", "35-44", "45-54", "55+"]},
            {"key": "tier", "label": "Loyalty tier", "type": "multi",
             "operators": OP_MULTI,
             "options": ["bronze", "silver", "gold", "platinum"]},
            {"key": "preferred_language", "label": "Preferred language",
             "type": "multi_async", "operators": OP_MULTI,
             "facet": "customers.preferred_language"},
            {"key": "source", "label": "Acquisition source", "type": "multi_async",
             "operators": OP_MULTI, "facet": "customers.source"},
            {"key": "card_validity", "label": "Card validity", "type": "multi",
             "operators": OP_MULTI, "options": ["Active", "Expired"]},
            {"key": "birthday_window", "label": "Birthday in next N days",
             "type": "number", "operators": ["lte"], "min": 0, "max": 365},
            {"key": "anniversary_window", "label": "Anniversary in next N days",
             "type": "number", "operators": ["lte"], "min": 0, "max": 365},
        ],
    },
    # ---------- 📞 Channel & Consent ----------
    "consent": {
        "label": "Channel & Consent",
        "icon": "MessageCircle",
        "fields": [
            {"key": "has_mobile", "label": "Has mobile number", "type": "boolean",
             "operators": OP_BOOL},
            {"key": "has_email", "label": "Has email address", "type": "boolean",
             "operators": OP_BOOL},
            {"key": "wa_opt_in", "label": "WhatsApp opt-in",
             "type": "boolean", "operators": OP_BOOL,
             "hint": "Defaults to true unless explicitly opted out"},
            {"key": "sms_opt_in", "label": "SMS opt-in", "type": "boolean",
             "operators": OP_BOOL},
            {"key": "email_opt_in", "label": "Email opt-in", "type": "boolean",
             "operators": OP_BOOL},
        ],
    },
    # ---------- 💰 Purchase Behaviour ----------
    "purchase": {
        "label": "Purchase Behaviour",
        "icon": "TrendingUp",
        "fields": [
            {"key": "lifecycle", "label": "Lifecycle bucket (R3)", "type": "multi",
             "operators": OP_MULTI,
             "options": ["never", "one_timer", "repeat", "frequent", "power"],
             "hint": "never=0 / one_timer=1 / repeat=2+ / frequent=3+ / power=6+ bills"},
            {"key": "visit_count", "label": "Lifetime visit count",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 1000},
            {"key": "lifetime_spend", "label": "Lifetime spend ₹",
             "type": "currency", "operators": OP_NUMERIC, "min": 0, "max": 10000000},
            {"key": "aov", "label": "Avg order value ₹ (lifetime_spend / visit_count)",
             "type": "currency", "operators": OP_NUMERIC, "min": 0, "max": 1000000},
            {"key": "recency_band", "label": "Last visit recency",
             "type": "multi", "operators": OP_MULTI,
             "options": ["never", "0-30d", "31-60d", "61-90d", "91-180d", "180d+"]},
            {"key": "days_since_last_visit", "label": "Days since last visit",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 3650},
            {"key": "categories_purchased", "label": "Categories purchased",
             "type": "multi_async", "operators": OP_MULTI,
             "facet": "transactions.items.category",
             "hint": "Customer has bought at least one item in these categories"},
            {"key": "skus_purchased", "label": "SKUs purchased", "type": "multi_async",
             "operators": OP_MULTI, "facet": "transactions.items.sku"},
            {"key": "distinct_sku_count", "label": "Distinct SKUs purchased",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 1000},
            {"key": "shopped_at_stores", "label": "Visited stores (any bill)",
             "type": "multi_async", "operators": OP_MULTI, "facet": "stores"},
        ],
    },
    # ---------- 🗓 Time-Window Behaviour ----------
    "time_window": {
        "label": "Time-Window Behaviour",
        "icon": "Calendar",
        "fields": [
            {"key": "first_purchase_at", "label": "First purchase date (R1)",
             "type": "date", "operators": OP_DATE},
            {"key": "last_visit_at", "label": "Last visit date",
             "type": "date", "operators": OP_DATE},
            {"key": "txn_count_in_window", "label": "# bills in date range",
             "type": "windowed_count", "operators": OP_NUMERIC,
             "hint": "Counts loyalty bills between window.from and window.to"},
            {"key": "day_pattern", "label": "Day-of-week pattern", "type": "multi",
             "operators": OP_MULTI,
             "options": ["weekday_only", "weekend_only", "mixed"]},
            {"key": "time_of_day_pattern", "label": "Time-of-day pattern",
             "type": "multi", "operators": OP_MULTI,
             "options": ["morning", "afternoon", "evening", "night"],
             "hint": "morning=06-12, afternoon=12-17, evening=17-21, night=21-06"},
        ],
    },
    # ---------- 🎁 Loyalty & Rewards ----------
    "loyalty": {
        "label": "Loyalty & Rewards",
        "icon": "Award",
        "fields": [
            {"key": "points_balance", "label": "Active points balance",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 1000000},
            {"key": "lifetime_points_earned", "label": "Lifetime points earned",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 10000000},
            {"key": "lifetime_points_redeemed", "label": "Lifetime points redeemed",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 10000000},
            {"key": "burn_ratio", "label": "Burn ratio (redeemed / earned)",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 100,
             "hint": "Percentage. 0 = never redeemed; 100 = fully burned"},
            {"key": "has_unredeemed_coupon", "label": "Has unredeemed coupon",
             "type": "boolean", "operators": OP_BOOL},
            {"key": "redeemed_last_n_days", "label": "Redeemed a reward in last N days",
             "type": "number", "operators": ["lte"], "min": 1, "max": 730},
        ],
    },
    # ---------- 🤝 Engagement & Risk ----------
    "engagement": {
        "label": "Engagement & Risk",
        "icon": "MessageCircle",
        "fields": [
            {"key": "churn_risk", "label": "Churn risk band", "type": "multi",
             "operators": OP_MULTI, "options": ["low", "medium", "high"]},
            {"key": "nps_band", "label": "NPS band (last response)", "type": "multi",
             "operators": OP_MULTI, "options": ["promoter", "passive", "detractor", "no_response"]},
            {"key": "nps_score", "label": "Last NPS score", "type": "number",
             "operators": OP_NUMERIC, "min": 0, "max": 10},
            {"key": "open_tickets", "label": "# open support tickets",
             "type": "number", "operators": OP_NUMERIC, "min": 0, "max": 50},
            {"key": "last_campaign_engagement", "label": "Last campaign engagement",
             "type": "multi", "operators": OP_MULTI,
             "options": ["opened", "clicked", "converted", "no_response"]},
            {"key": "campaign_cooldown_days", "label": "Exclude if messaged in last N days",
             "type": "number", "operators": ["gte"], "min": 1, "max": 365,
             "hint": "Drops customers contacted by ANY campaign in this window"},
        ],
    },
}


# ============================================================
# Validation
# ============================================================
class Rule(BaseModel):
    field: str
    operator: str
    value: Any = None


class FilterGroup(BaseModel):
    op: str = "AND"  # AND | OR
    rules: List[Any] = Field(default_factory=list)  # Rule or FilterGroup


class AudienceIn(BaseModel):
    tree: Dict[str, Any]
    window: Optional[Dict[str, str]] = None
    page: int = 1
    page_size: int = 25
    sort_by: str = "lifetime_spend"  # or last_visit_at, visit_count, points_balance
    sort_dir: int = -1  # -1 desc, 1 asc


class AudienceExportIn(BaseModel):
    tree: Dict[str, Any]
    window: Optional[Dict[str, str]] = None
    sort_by: str = "lifetime_spend"
    sort_dir: int = -1
    format: str = "csv"  # csv | xlsx | pdf
    segment_name: Optional[str] = None
    max_rows: int = 200000  # hard cap


class PreviewIn(BaseModel):
    tree: Dict[str, Any]
    # Optional global windows applied to "windowed_count" fields
    window: Optional[Dict[str, str]] = None


class FacetsIn(BaseModel):
    source: str  # e.g. "customers.city" or "stores"
    query: Optional[str] = None
    limit: int = 30


class SegmentIn(BaseModel):
    name: str
    description: Optional[str] = ""
    tree: Dict[str, Any]
    window: Optional[Dict[str, str]] = None


def _build_field_index() -> Dict[str, Dict[str, Any]]:
    """Flatten the FILTER_SCHEMA into a lookup by field key."""
    idx = {}
    for cat in FILTER_SCHEMA.values():
        for f in cat["fields"]:
            idx[f["key"]] = f
    return idx


FIELD_INDEX = _build_field_index()


# ============================================================
# Compiler — translates a validated filter tree to a Mongo match dict
# ============================================================
async def _resolve_transaction_filter_to_mobiles(field: str, op: str, value: Any) -> List[str]:
    """For txn-derived filters (categories/SKUs purchased, day-pattern, time-of-day,
    visited stores, distinct_sku_count, last campaign engagement, etc.), find the
    set of customer mobiles that satisfy and return the list."""
    if field == "skus_purchased":
        match_value = value if isinstance(value, list) else [value]
        if op == "in":
            mobiles = await transactions_col.distinct(
                "customer_mobile",
                {"customer_mobile": {"$nin": [None, ""]},
                 "items.sku": {"$in": match_value}},
            )
            return [m for m in mobiles if m]
        if op == "not_in":
            # Customers WITHOUT any of these SKUs — anti-set
            all_mob = await transactions_col.distinct(
                "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}}
            )
            with_them = await transactions_col.distinct(
                "customer_mobile",
                {"customer_mobile": {"$nin": [None, ""]}, "items.sku": {"$in": match_value}},
            )
            wset = set(m for m in with_them if m)
            return [m for m in all_mob if m and m not in wset]

    if field == "categories_purchased":
        match_value = value if isinstance(value, list) else [value]
        if op == "in":
            mobiles = await transactions_col.distinct(
                "customer_mobile",
                {"customer_mobile": {"$nin": [None, ""]},
                 "items.category": {"$in": match_value}},
            )
            return [m for m in mobiles if m]
        if op == "not_in":
            all_mob = await transactions_col.distinct(
                "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}}
            )
            with_them = await transactions_col.distinct(
                "customer_mobile",
                {"customer_mobile": {"$nin": [None, ""]},
                 "items.category": {"$in": match_value}},
            )
            wset = set(m for m in with_them if m)
            return [m for m in all_mob if m and m not in wset]

    if field == "shopped_at_stores":
        match_value = value if isinstance(value, list) else [value]
        if op == "in":
            mobiles = await transactions_col.distinct(
                "customer_mobile",
                {"customer_mobile": {"$nin": [None, ""]}, "store_id": {"$in": match_value}},
            )
            return [m for m in mobiles if m]
        if op == "not_in":
            all_mob = await transactions_col.distinct(
                "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}}
            )
            with_them = await transactions_col.distinct(
                "customer_mobile",
                {"customer_mobile": {"$nin": [None, ""]}, "store_id": {"$in": match_value}},
            )
            wset = set(m for m in with_them if m)
            return [m for m in all_mob if m and m not in wset]

    if field == "distinct_sku_count":
        # Need to aggregate: count distinct SKUs per mobile
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
            {"$unwind": {"path": "$items", "preserveNullAndEmptyArrays": False}},
            {"$group": {"_id": "$customer_mobile", "skus": {"$addToSet": "$items.sku"}}},
            {"$project": {"n": {"$size": "$skus"}}},
        ]
        results = await transactions_col.aggregate(pipe).to_list(200000)
        mobiles = []
        for r in results:
            n = r["n"]
            if _matches_numeric(n, op, value):
                mobiles.append(r["_id"])
        return mobiles

    if field == "day_pattern":
        # Determine each customer's dominant pattern
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
            {"$project": {
                "customer_mobile": 1,
                "dow": {"$dayOfWeek": {"$dateFromString": {"dateString": "$bill_date"}}},
            }},
            {"$group": {
                "_id": "$customer_mobile",
                "weekday_count": {"$sum": {"$cond": [{"$and": [{"$gte": ["$dow", 2]}, {"$lte": ["$dow", 6]}]}, 1, 0]}},
                "weekend_count": {"$sum": {"$cond": [{"$or": [{"$eq": ["$dow", 1]}, {"$eq": ["$dow", 7]}]}, 1, 0]}},
            }},
        ]
        rows = await transactions_col.aggregate(pipe).to_list(200000)
        targets = value if isinstance(value, list) else [value]
        out = []
        for r in rows:
            wkd, wkn = r["weekday_count"], r["weekend_count"]
            if wkd == 0 and wkn == 0:
                continue
            pattern = ("weekday_only" if wkn == 0 else "weekend_only" if wkd == 0 else "mixed")
            if op == "in" and pattern in targets:
                out.append(r["_id"])
            elif op == "not_in" and pattern not in targets:
                out.append(r["_id"])
        return out

    if field == "time_of_day_pattern":
        # Determine each customer's most-frequent time bucket
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
            {"$project": {
                "customer_mobile": 1,
                "hour": {"$hour": {"$dateFromString": {"dateString": "$bill_date"}}},
            }},
            {"$project": {
                "customer_mobile": 1,
                "bucket": {
                    "$switch": {
                        "branches": [
                            {"case": {"$and": [{"$gte": ["$hour", 6]}, {"$lt": ["$hour", 12]}]}, "then": "morning"},
                            {"case": {"$and": [{"$gte": ["$hour", 12]}, {"$lt": ["$hour", 17]}]}, "then": "afternoon"},
                            {"case": {"$and": [{"$gte": ["$hour", 17]}, {"$lt": ["$hour", 21]}]}, "then": "evening"},
                        ],
                        "default": "night",
                    },
                },
            }},
            {"$group": {"_id": {"mob": "$customer_mobile", "bucket": "$bucket"}, "n": {"$sum": 1}}},
            {"$sort": {"n": -1}},
            {"$group": {"_id": "$_id.mob", "bucket": {"$first": "$_id.bucket"}}},
        ]
        rows = await transactions_col.aggregate(pipe).to_list(200000)
        targets = value if isinstance(value, list) else [value]
        if op == "in":
            return [r["_id"] for r in rows if r["bucket"] in targets]
        if op == "not_in":
            return [r["_id"] for r in rows if r["bucket"] not in targets]

    if field == "last_campaign_engagement":
        # Join message_log: per-mobile latest event status
        targets = value if isinstance(value, list) else [value]
        pipe = [
            {"$match": {"to": {"$nin": [None, ""]}}},
            {"$sort": {"created_at": -1}},
            {"$group": {"_id": "$to", "status": {"$first": "$status"}}},
        ]
        rows = await message_log_col.aggregate(pipe).to_list(200000)
        # Map message log status → engagement label
        def to_label(s: str) -> str:
            s = (s or "").lower()
            if "convert" in s:
                return "converted"
            if "click" in s:
                return "clicked"
            if "open" in s or "delivered" in s:
                return "opened"
            return "no_response"
        if op == "in":
            return [r["_id"] for r in rows if to_label(r["status"]) in targets]
        if op == "not_in":
            return [r["_id"] for r in rows if to_label(r["status"]) not in targets]

    if field == "campaign_cooldown_days":
        # Drop customers messaged in last N days (operator interpretation: gte = cooldown N+ days = exclude messaged ≤ N)
        days = int(value or 0)
        if days <= 0:
            return []
        cutoff = _iso(_now() - timedelta(days=days))
        recently_messaged = await message_log_col.distinct(
            "to", {"created_at": {"$gte": cutoff}, "to": {"$nin": [None, ""]}}
        )
        # We return the INVERTED set: customers NOT in recently_messaged.
        # But to keep compiler uniform we return mobiles to EXCLUDE; caller will use $nin.
        return list(set(m for m in recently_messaged if m))  # consumer must handle as exclusion

    if field == "txn_count_in_window":
        # Count bills within the window for each mobile; match operator
        # window is expected on the global PreviewIn.window — surfaced via closure
        return []  # handled at compile level using _resolve_windowed_count_to_mobiles

    if field in ("has_unredeemed_coupon",):
        # Look up coupon_redemptions: a customer has unredeemed coupon if any active coupon
        # is assigned but not yet redeemed. Simplified: customers WITHOUT any redemption row in last 90d.
        bool_val = bool(value)
        all_mob = await customers_col.distinct("mobile", {"mobile": {"$nin": [None, ""]}})
        redeemed_mob = await coupon_redemptions_col.distinct(
            "customer_mobile", {"customer_mobile": {"$nin": [None, ""]}}
        )
        rs = set(m for m in redeemed_mob if m)
        with_unredeemed = [m for m in all_mob if m not in rs]
        if bool_val:
            return with_unredeemed
        return list(rs)

    if field == "redeemed_last_n_days":
        days = int(value or 0)
        if days <= 0:
            return []
        cutoff = _iso(_now() - timedelta(days=days))
        mobiles = await coupon_redemptions_col.distinct(
            "customer_mobile",
            {"created_at": {"$gte": cutoff}, "customer_mobile": {"$nin": [None, ""]}},
        )
        return [m for m in mobiles if m]

    if field == "open_tickets":
        # Aggregate per customer mobile
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]},
                        "status": {"$in": ["open", "in_progress", "escalated"]}}},
            {"$group": {"_id": "$customer_mobile", "n": {"$sum": 1}}},
        ]
        rows = await tickets_col.aggregate(pipe).to_list(200000)
        return [r["_id"] for r in rows if _matches_numeric(r["n"], op, value)]

    if field == "nps_band":
        # Per-customer last NPS response → band
        targets = value if isinstance(value, list) else [value]
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
            {"$sort": {"created_at": -1}},
            {"$group": {"_id": "$customer_mobile", "score": {"$first": "$score"}}},
        ]
        rows = await nps_col.aggregate(pipe).to_list(200000)
        def band(s: int) -> str:
            if s is None:
                return "no_response"
            if s >= 9:
                return "promoter"
            if s >= 7:
                return "passive"
            return "detractor"
        if op == "in":
            return [r["_id"] for r in rows if band(r["score"]) in targets]
        if op == "not_in":
            return [r["_id"] for r in rows if band(r["score"]) not in targets]

    if field == "nps_score":
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
            {"$sort": {"created_at": -1}},
            {"$group": {"_id": "$customer_mobile", "score": {"$first": "$score"}}},
        ]
        rows = await nps_col.aggregate(pipe).to_list(200000)
        return [r["_id"] for r in rows if _matches_numeric(r["score"], op, value)]

    return []


def _matches_numeric(n: float, op: str, value: Any) -> bool:
    try:
        if op == "eq":
            return n == float(value)
        if op == "neq":
            return n != float(value)
        if op == "gte":
            return n >= float(value)
        if op == "lte":
            return n <= float(value)
        if op == "between":
            a, b = float(value[0]), float(value[1])
            return a <= n <= b
    except (TypeError, ValueError):
        return False
    return False


async def _resolve_windowed_count_to_mobiles(op: str, value: Any, window: Optional[Dict[str, str]]) -> List[str]:
    if not window or not window.get("from"):
        return []
    start = window["from"]
    end = window.get("to") or _iso(_now())
    pipe = [
        {"$match": {"customer_mobile": {"$nin": [None, ""]},
                    "bill_date": {"$gte": start, "$lte": end}}},
        {"$group": {"_id": "$customer_mobile", "n": {"$sum": 1}}},
    ]
    rows = await transactions_col.aggregate(pipe).to_list(200000)
    return [r["_id"] for r in rows if _matches_numeric(r["n"], op, value)]


def _compile_rule_simple(field_meta: Dict[str, Any], op: str, value: Any) -> Optional[Dict[str, Any]]:
    """Translate a rule to a Mongo filter clause on the customers collection.
    Returns None when the field is transaction-derived (caller will use the
    mobile-list path)."""
    key = field_meta["key"]

    # ---- Direct-mapped scalar fields ----
    if key in {"city", "state", "country_code", "preferred_language", "source", "tier",
                "gender", "card_validity", "home_store_id", "churn_risk"}:
        if op == "in":
            return {key: {"$in": value if isinstance(value, list) else [value]}}
        if op == "not_in":
            return {key: {"$nin": value if isinstance(value, list) else [value]}}

    if key in {"visit_count", "lifetime_spend", "lifetime_points_earned",
                "lifetime_points_redeemed", "points_balance"}:
        return _numeric_to_clause(key, op, value)

    if key == "aov":
        # Computed virtual via $expr; we use $expr-equivalent: lifetime_spend / max(visit_count,1)
        # Translate to lifetime_spend bounds approximation (over-fetch then filter at app layer)
        # For simplicity here, return $expr with $divide.
        v = value
        if op == "between":
            lo, hi = float(v[0]), float(v[1])
            expr = {"$and": [
                {"$gte": [{"$divide": ["$lifetime_spend", {"$max": ["$visit_count", 1]}]}, lo]},
                {"$lte": [{"$divide": ["$lifetime_spend", {"$max": ["$visit_count", 1]}]}, hi]},
            ]}
        elif op == "gte":
            expr = {"$gte": [{"$divide": ["$lifetime_spend", {"$max": ["$visit_count", 1]}]}, float(v)]}
        elif op == "lte":
            expr = {"$lte": [{"$divide": ["$lifetime_spend", {"$max": ["$visit_count", 1]}]}, float(v)]}
        elif op == "eq":
            expr = {"$eq": [{"$divide": ["$lifetime_spend", {"$max": ["$visit_count", 1]}]}, float(v)]}
        elif op == "neq":
            expr = {"$ne": [{"$divide": ["$lifetime_spend", {"$max": ["$visit_count", 1]}]}, float(v)]}
        else:
            return None
        return {"$expr": expr}

    if key == "burn_ratio":
        v = value
        # burn_ratio = lifetime_points_redeemed / max(lifetime_points_earned,1) * 100
        ratio_expr = {"$multiply": [
            {"$divide": ["$lifetime_points_redeemed", {"$max": ["$lifetime_points_earned", 1]}]},
            100,
        ]}
        if op == "between":
            lo, hi = float(v[0]), float(v[1])
            return {"$expr": {"$and": [{"$gte": [ratio_expr, lo]}, {"$lte": [ratio_expr, hi]}]}}
        if op == "gte":
            return {"$expr": {"$gte": [ratio_expr, float(v)]}}
        if op == "lte":
            return {"$expr": {"$lte": [ratio_expr, float(v)]}}
        return None

    if key == "lifecycle":
        # one_timer = visit_count == 1, repeat >=2, frequent >=3, power >=6, never = 0
        targets = value if isinstance(value, list) else [value]
        clauses = []
        for t in targets:
            if t == "never":
                clauses.append({"$or": [{"visit_count": 0}, {"visit_count": None}, {"visit_count": {"$exists": False}}]})
            elif t == "one_timer":
                clauses.append({"visit_count": 1})
            elif t == "repeat":
                clauses.append({"visit_count": {"$gte": 2, "$lt": 3}})
            elif t == "frequent":
                clauses.append({"visit_count": {"$gte": 3, "$lt": 6}})
            elif t == "power":
                clauses.append({"visit_count": {"$gte": 6}})
        if not clauses:
            return None
        if op == "in":
            return {"$or": clauses}
        if op == "not_in":
            return {"$nor": clauses}

    if key == "recency_band":
        targets = value if isinstance(value, list) else [value]
        now = _now()
        bands_map = {
            "0-30d": (now - timedelta(days=30), now),
            "31-60d": (now - timedelta(days=60), now - timedelta(days=30)),
            "61-90d": (now - timedelta(days=90), now - timedelta(days=60)),
            "91-180d": (now - timedelta(days=180), now - timedelta(days=90)),
            "180d+": (None, now - timedelta(days=180)),
            "never": None,
        }
        clauses = []
        for t in targets:
            if t == "never":
                clauses.append({"$or": [
                    {"last_visit_at": None},
                    {"last_visit_at": {"$exists": False}},
                ]})
            elif t in bands_map and bands_map[t] is not None:
                lo, hi = bands_map[t]
                c: Dict[str, Any] = {}
                if lo is not None and hi is not None:
                    c["last_visit_at"] = {"$gte": _iso(lo), "$lt": _iso(hi)}
                elif hi is not None:
                    c["last_visit_at"] = {"$lt": _iso(hi)}
                clauses.append(c)
        if not clauses:
            return None
        if op == "in":
            return {"$or": clauses}
        if op == "not_in":
            return {"$nor": clauses}

    if key == "days_since_last_visit":
        # Translate to last_visit_at bounds (inverted: bigger days = older last_visit_at)
        now = _now()
        v = value
        if op == "between":
            d_lo, d_hi = int(v[0]), int(v[1])
            return {"last_visit_at": {
                "$gte": _iso(now - timedelta(days=d_hi)),
                "$lte": _iso(now - timedelta(days=d_lo)),
            }}
        if op == "gte":
            return {"last_visit_at": {"$lte": _iso(now - timedelta(days=int(v)))}}
        if op == "lte":
            return {"last_visit_at": {"$gte": _iso(now - timedelta(days=int(v)))}}
        if op == "eq":
            target = now - timedelta(days=int(v))
            return {"last_visit_at": {"$gte": _iso(target - timedelta(hours=12)),
                                        "$lt": _iso(target + timedelta(hours=12))}}
        return None

    if key == "first_purchase_at":
        return _date_to_clause(key, op, value)

    if key == "last_visit_at":
        return _date_to_clause(key, op, value)

    if key == "age_band":
        targets = value if isinstance(value, list) else [value]
        now = _now()
        bands = {
            "18-24": (24, 18),
            "25-34": (34, 25),
            "35-44": (44, 35),
            "45-54": (54, 45),
            "55+": (200, 55),
        }
        clauses = []
        for t in targets:
            if t in bands:
                upper_age, lower_age = bands[t]
                # birthday between now - upper_age years and now - lower_age years
                hi = now.replace(year=now.year - lower_age)
                lo = now.replace(year=now.year - upper_age - 1)
                clauses.append({"birthday": {"$gte": _iso(lo), "$lte": _iso(hi)}})
        if not clauses:
            return None
        if op == "in":
            return {"$or": clauses}
        if op == "not_in":
            return {"$nor": clauses}

    if key == "birthday_window":
        # Customers whose next birthday falls within N days. Compare month-day.
        n = int(value or 0)
        if n <= 0:
            return None
        # Compute date range as MM-DD strings; birthday stored as ISO. Use $expr.
        now = _now()
        target = now + timedelta(days=n)
        # If wrap-around (Dec→Jan), split into two ranges
        mm_now = f"{now.month:02d}-{now.day:02d}"
        mm_target = f"{target.month:02d}-{target.day:02d}"
        # Use substring slice of ISO date (positions 5-10 = "MM-DD")
        if mm_now <= mm_target:
            return {"$expr": {"$and": [
                {"$gte": [{"$substr": ["$birthday", 5, 5]}, mm_now]},
                {"$lte": [{"$substr": ["$birthday", 5, 5]}, mm_target]},
            ]}}
        return {"$expr": {"$or": [
            {"$gte": [{"$substr": ["$birthday", 5, 5]}, mm_now]},
            {"$lte": [{"$substr": ["$birthday", 5, 5]}, mm_target]},
        ]}}

    if key == "anniversary_window":
        n = int(value or 0)
        if n <= 0:
            return None
        now = _now()
        target = now + timedelta(days=n)
        mm_now = f"{now.month:02d}-{now.day:02d}"
        mm_target = f"{target.month:02d}-{target.day:02d}"
        if mm_now <= mm_target:
            return {"$expr": {"$and": [
                {"$gte": [{"$substr": ["$anniversary", 5, 5]}, mm_now]},
                {"$lte": [{"$substr": ["$anniversary", 5, 5]}, mm_target]},
            ]}}
        return {"$expr": {"$or": [
            {"$gte": [{"$substr": ["$anniversary", 5, 5]}, mm_now]},
            {"$lte": [{"$substr": ["$anniversary", 5, 5]}, mm_target]},
        ]}}

    if key == "has_mobile":
        return {"mobile": {"$nin": [None, ""]}} if value else {"mobile": {"$in": [None, ""]}}
    if key == "has_email":
        return {"email": {"$nin": [None, ""]}} if value else {"$or": [{"email": {"$in": [None, ""]}}, {"email": {"$exists": False}}]}
    if key == "wa_opt_in":
        # Treat missing as opted-in (industry default)
        return {"wa_opt_in": {"$ne": False}} if value else {"wa_opt_in": False}
    if key == "sms_opt_in":
        return {"sms_opt_in": {"$ne": False}} if value else {"sms_opt_in": False}
    if key == "email_opt_in":
        return {"email_opt_in": {"$ne": False}} if value else {"email_opt_in": False}

    # If we reach here, this field is transaction-derived (return None → handled via mobile-list path)
    return None


def _numeric_to_clause(field: str, op: str, value: Any) -> Optional[Dict[str, Any]]:
    if op == "between":
        return {field: {"$gte": float(value[0]), "$lte": float(value[1])}}
    if op == "gte":
        return {field: {"$gte": float(value)}}
    if op == "lte":
        return {field: {"$lte": float(value)}}
    if op == "eq":
        return {field: float(value)}
    if op == "neq":
        return {field: {"$ne": float(value)}}
    return None


def _date_to_clause(field: str, op: str, value: Any) -> Optional[Dict[str, Any]]:
    if op == "between":
        return {field: {"$gte": value[0], "$lte": value[1]}}
    if op == "gte":
        return {field: {"$gte": value}}
    if op == "lte":
        return {field: {"$lte": value}}
    return None


TXN_DERIVED_FIELDS = {
    "skus_purchased", "categories_purchased", "shopped_at_stores",
    "distinct_sku_count", "day_pattern", "time_of_day_pattern",
    "last_campaign_engagement", "campaign_cooldown_days",
    "has_unredeemed_coupon", "redeemed_last_n_days",
    "open_tickets", "nps_band", "nps_score", "txn_count_in_window",
}

# fields that need lookup via stores (city / region of home store)
HOME_STORE_DERIVED_FIELDS = {"home_store_region", "home_store_city"}


async def _resolve_home_store_derived(field: str, op: str, value: Any) -> List[str]:
    """Resolve home_store_city/region to a list of customer mobiles."""
    if field == "home_store_region":
        col = "region"
    elif field == "home_store_city":
        col = "city"
    else:
        return []
    val = value if isinstance(value, list) else [value]
    if op == "in":
        store_ids = await stores_col.distinct("id", {col: {"$in": val}})
    elif op == "not_in":
        store_ids = await stores_col.distinct("id", {col: {"$nin": val}})
    else:
        return []
    mobiles = await customers_col.distinct(
        "mobile", {"home_store_id": {"$in": store_ids}, "mobile": {"$nin": [None, ""]}}
    )
    return [m for m in mobiles if m]


async def compile_tree(tree: Dict[str, Any], window: Optional[Dict[str, str]] = None,
                       _depth: int = 0) -> Dict[str, Any]:
    """Translate a filter tree into a customer-collection match dict.

    Accepts either a group ({op, rules}) or a single rule
    ({field, operator, value}) at the root.
    """
    if _depth > 2:
        raise HTTPException(400, "Max 2 levels of nesting allowed")

    # Allow callers to pass a bare rule at the root — wrap it in an AND group.
    if tree and "field" in tree and "operator" in tree and "rules" not in tree:
        tree = {"op": "AND", "rules": [tree]}

    op = (tree.get("op") or "AND").upper()
    if op not in ("AND", "OR"):
        raise HTTPException(400, f"Invalid op '{op}', must be AND or OR")

    rules = tree.get("rules") or []
    if not rules:
        return {}

    # Translate each rule
    clauses: List[Dict[str, Any]] = []
    exclude_mobiles: List[str] = []  # cooldown drops; always AND-applied
    for raw in rules:
        if isinstance(raw, dict) and "op" in raw and "rules" in raw:
            sub = await compile_tree(raw, window, _depth + 1)
            if sub:
                clauses.append(sub)
            continue
        if not isinstance(raw, dict):
            continue
        field = raw.get("field")
        operator = raw.get("operator")
        value = raw.get("value")
        if not field or not operator:
            continue
        fmeta = FIELD_INDEX.get(field)
        if not fmeta:
            continue

        # Transaction-derived path: resolve to mobile list then $in
        if field in TXN_DERIVED_FIELDS:
            if field == "txn_count_in_window":
                mobiles = await _resolve_windowed_count_to_mobiles(operator, value, window)
            else:
                mobiles = await _resolve_transaction_filter_to_mobiles(field, operator, value)
            if field == "campaign_cooldown_days":
                # exclusion list — handled outside
                exclude_mobiles = list(set(exclude_mobiles + mobiles))
                continue
            clauses.append({"mobile": {"$in": mobiles}})
            continue

        # Home-store-derived path
        if field in HOME_STORE_DERIVED_FIELDS:
            mobiles = await _resolve_home_store_derived(field, operator, value)
            clauses.append({"mobile": {"$in": mobiles}})
            continue

        # Direct path
        compiled = _compile_rule_simple(fmeta, operator, value)
        if compiled is not None:
            clauses.append(compiled)

    if not clauses and not exclude_mobiles:
        return {}

    if op == "AND":
        final = {"$and": clauses} if clauses else {}
    else:
        final = {"$or": clauses} if clauses else {}

    # Cooldown exclusion always wins (AND-applied at the top level only)
    if exclude_mobiles and _depth == 0:
        cooldown_clause = {"mobile": {"$nin": exclude_mobiles}}
        if final:
            final = {"$and": [final, cooldown_clause]}
        else:
            final = cooldown_clause

    return final


# ============================================================
# Route handlers
# ============================================================
@router.get("/filter-schema")
async def get_filter_schema(user: dict = Depends(get_current_user)):
    """Return the full filter taxonomy for the frontend to render."""
    return {"schema": FILTER_SCHEMA}


@router.post("/facets")
async def get_facets(body: FacetsIn, user: dict = Depends(get_current_user)):
    """Return distinct values for type-ahead dropdowns.

    `source` can be:
      - "customers.<field>" — distinct on customers collection
      - "transactions.items.<field>" — distinct on transactions items
      - "stores" — full store list
      - "stores.<field>" — distinct on stores
    """
    src = body.source or ""
    q = (body.query or "").strip()
    limit = max(1, min(body.limit or 30, 200))

    if src == "stores":
        match = {}
        if q:
            match["$or"] = [
                {"name": {"$regex": re.escape(q), "$options": "i"}},
                {"code": {"$regex": re.escape(q), "$options": "i"}},
                {"city": {"$regex": re.escape(q), "$options": "i"}},
            ]
        rows = await stores_col.find(match, {"_id": 0, "id": 1, "name": 1, "code": 1, "city": 1, "region": 1}).limit(limit).to_list(limit)
        return {"values": [{"value": r["id"], "label": f"{r['name']} · {r.get('code','')} · {r.get('city','')}"} for r in rows]}

    if src.startswith("stores."):
        field = src.split(".", 1)[1]
        match = {field: {"$nin": [None, ""]}}
        if q:
            match[field] = {"$regex": re.escape(q), "$options": "i", "$nin": [None, ""]}
        vals = await stores_col.distinct(field, match)
        return {"values": [{"value": v, "label": v} for v in sorted([x for x in vals if x])[:limit]]}

    if src.startswith("customers."):
        field = src.split(".", 1)[1]
        match = {field: {"$nin": [None, ""]}}
        if q:
            match[field] = {"$regex": re.escape(q), "$options": "i", "$nin": [None, ""]}
        vals = await customers_col.distinct(field, match)
        return {"values": [{"value": v, "label": str(v)} for v in sorted([x for x in vals if x])[:limit]]}

    if src.startswith("transactions.items."):
        field = src.split(".", 2)[2]
        pipe = [
            {"$match": {"customer_mobile": {"$nin": [None, ""]}}},
            {"$unwind": "$items"},
            {"$group": {"_id": f"$items.{field}"}},
            {"$match": {"_id": {"$nin": [None, ""]}}},
        ]
        if q:
            pipe.append({"$match": {"_id": {"$regex": re.escape(q), "$options": "i"}}})
        pipe.append({"$limit": limit * 2})
        rows = await transactions_col.aggregate(pipe).to_list(limit * 2)
        vals = sorted([r["_id"] for r in rows if r["_id"]])[:limit]
        return {"values": [{"value": v, "label": str(v)} for v in vals]}

    return {"values": []}


async def _compute_preview(tree: Dict[str, Any], window: Optional[Dict[str, str]]) -> Dict[str, Any]:
    match = await compile_tree(tree, window)
    total = await customers_col.count_documents(match)

    # Reach by channel
    wa_match = {**match, "mobile": {"$nin": [None, ""]}, "wa_opt_in": {"$ne": False}}
    sms_match = {**match, "mobile": {"$nin": [None, ""]}, "sms_opt_in": {"$ne": False}}
    email_match = {**match, "email": {"$nin": [None, ""]}, "email_opt_in": {"$ne": False}}
    opted_out_match = {**match, "$or": [
        {"wa_opt_in": False}, {"sms_opt_in": False}, {"email_opt_in": False}
    ]}

    # Sample 5 customers for sanity
    sample = await customers_col.find(
        match,
        {"_id": 0, "id": 1, "name": 1, "mobile": 1, "city": 1, "tier": 1,
         "lifetime_spend": 1, "visit_count": 1, "first_purchase_at": 1, "last_visit_at": 1}
    ).limit(5).to_list(5)

    return {
        "matched_total": total,
        "reach": {
            "whatsapp": await customers_col.count_documents(wa_match),
            "sms": await customers_col.count_documents(sms_match),
            "email": await customers_col.count_documents(email_match),
            "opted_out": await customers_col.count_documents(opted_out_match),
        },
        "sample": sample,
        "compiled_filter": match,
    }


@router.post("/preview")
async def preview(body: PreviewIn, user: dict = Depends(get_current_user)):
    """Live preview — count + reach breakdown + 5 sample customers."""
    return await _compute_preview(body.tree, body.window)


@router.post("/audience")
async def audience(body: AudienceIn, user: dict = Depends(get_current_user)):
    """Rich paginated audience list for the cohort/filter editor.

    Returns full customer rows: mobile, name, city, tier, visit_count,
    lifetime_spend, last_visit_at, points_balance, home_store_id, etc.
    """
    match = await compile_tree(body.tree, body.window)
    page = max(1, body.page)
    page_size = max(1, min(body.page_size, 200))
    skip = (page - 1) * page_size

    total = await customers_col.count_documents(match)

    allowed_sort = {"lifetime_spend", "last_visit_at", "visit_count",
                     "points_balance", "first_purchase_at", "name"}
    sort_by = body.sort_by if body.sort_by in allowed_sort else "lifetime_spend"
    sort_dir = -1 if body.sort_dir == -1 else 1

    rows = await customers_col.find(
        match,
        {"_id": 0, "id": 1, "mobile": 1, "name": 1, "city": 1, "tier": 1,
         "visit_count": 1, "lifetime_spend": 1, "last_visit_at": 1,
         "first_purchase_at": 1, "points_balance": 1, "lifetime_points_earned": 1,
         "lifetime_points_redeemed": 1, "home_store_id": 1, "churn_risk": 1,
         "email": 1, "gender": 1, "birthday": 1}
    ).sort(sort_by, sort_dir).skip(skip).limit(page_size).to_list(page_size)

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": (total + page_size - 1) // page_size if page_size else 1,
        "rows": rows,
    }


# ============================================================
# Audience Export — full report in CSV / XLSX / PDF
# ============================================================
EXPORT_COLUMNS = [
    ("mobile",                   "Mobile"),
    ("name",                     "Name"),
    ("email",                    "Email"),
    ("city",                     "City"),
    ("tier",                     "Tier"),
    ("gender",                   "Gender"),
    ("visit_count",              "Bills"),
    ("lifetime_spend",           "Lifetime Spend"),
    ("first_purchase_at",        "First Purchase"),
    ("last_visit_at",            "Last Visit"),
    ("points_balance",           "Points Balance"),
    ("lifetime_points_earned",   "Lifetime Earned"),
    ("lifetime_points_redeemed", "Lifetime Redeemed"),
    ("churn_risk",               "Churn Risk"),
    ("home_store_id",            "Home Store ID"),
    ("birthday",                 "Birthday"),
]

EXPORT_PROJECTION = {"_id": 0, **{k: 1 for k, _ in EXPORT_COLUMNS}}


def _fmt_export_cell(key: str, value: Any) -> str:
    if value is None or value == "":
        return ""
    if key in {"first_purchase_at", "last_visit_at"}:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        s = str(value)
        return s[:10] if len(s) >= 10 else s
    if key in {"lifetime_spend"}:
        try:
            return f"{float(value):.2f}"
        except (TypeError, ValueError):
            return str(value)
    return str(value)


@router.post("/audience/export")
async def audience_export(body: AudienceExportIn, user: dict = Depends(get_current_user)):
    """Export the FULL audience (not just current page) as CSV, XLSX, or PDF.

    - CSV streams row-by-row from MongoDB (constant memory).
    - XLSX uses openpyxl write_only workbook (low memory for large sheets).
    - PDF uses reportlab with a branded KAZO / Fundle header. For PDFs we cap
      the table at the first 2000 rows for readability and append a footer
      indicating how many rows were truncated.
    """
    fmt = (body.format or "csv").lower()
    if fmt not in {"csv", "xlsx", "pdf"}:
        raise HTTPException(400, "format must be csv, xlsx or pdf")

    match = await compile_tree(body.tree, body.window)
    allowed_sort = {"lifetime_spend", "last_visit_at", "visit_count",
                    "points_balance", "first_purchase_at", "name"}
    sort_by = body.sort_by if body.sort_by in allowed_sort else "lifetime_spend"
    sort_dir = -1 if body.sort_dir == -1 else 1
    max_rows = max(1, min(int(body.max_rows or 200000), 500000))

    total_matched = await customers_col.count_documents(match)
    seg_name = (body.segment_name or "audience").strip() or "audience"
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", seg_name)[:60] or "audience"
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    cursor = (
        customers_col.find(match, EXPORT_PROJECTION)
        .sort(sort_by, sort_dir)
        .limit(max_rows)
    )

    if fmt == "csv":
        async def stream_csv():
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow([label for _, label in EXPORT_COLUMNS])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)
            async for doc in cursor:
                writer.writerow([_fmt_export_cell(k, doc.get(k)) for k, _ in EXPORT_COLUMNS])
                if buf.tell() > 32 * 1024:
                    yield buf.getvalue()
                    buf.seek(0)
                    buf.truncate(0)
            if buf.tell() > 0:
                yield buf.getvalue()

        filename = f"{safe_name}_{ts}.csv"
        return StreamingResponse(
            stream_csv(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Audience")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="3B1A2A")  # KAZO burgundy
        # write_only doesn't support set_column widths easily; we set after
        from openpyxl.cell import WriteOnlyCell
        header_cells = []
        for _, label in EXPORT_COLUMNS:
            c = WriteOnlyCell(ws, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            header_cells.append(c)
        ws.append(header_cells)
        rows_written = 0
        async for doc in cursor:
            ws.append([_fmt_export_cell(k, doc.get(k)) for k, _ in EXPORT_COLUMNS])
            rows_written += 1
        # Column widths (write_only requires dimensions set explicitly)
        widths = [14, 22, 26, 16, 10, 8, 8, 16, 14, 14, 12, 14, 14, 12, 24, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        # Summary sheet
        meta = wb.create_sheet("Summary")
        meta.append(["KAZO · Fundle — Segment Audience Export"])
        meta.append(["Segment", seg_name])
        meta.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
        meta.append(["Generated by", user.get("email", "—")])
        meta.append(["Total matched (DB)", total_matched])
        meta.append(["Rows exported", rows_written])
        if rows_written < total_matched:
            meta.append(["Note", f"Truncated to first {rows_written} rows (cap {max_rows})."])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{safe_name}_{ts}.xlsx"
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # PDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )

    PDF_TABLE_CAP = 2000  # readable cap; CSV/XLSX hold the full dump
    pdf_cols = [
        ("mobile", "Mobile", 22),
        ("name", "Name", 38),
        ("city", "City", 22),
        ("tier", "Tier", 14),
        ("visit_count", "Bills", 12),
        ("lifetime_spend", "Spend", 22),
        ("last_visit_at", "Last Visit", 22),
        ("points_balance", "Points", 16),
    ]

    rows_collected: List[List[str]] = []
    total_exported = 0
    truncated_at = None
    async for doc in cursor:
        total_exported += 1
        if len(rows_collected) < PDF_TABLE_CAP:
            rows_collected.append([_fmt_export_cell(k, doc.get(k)) for k, _, _ in pdf_cols])
        elif truncated_at is None:
            truncated_at = PDF_TABLE_CAP

    buf = io.BytesIO()
    doc_pdf = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"KAZO Audience — {seg_name}",
    )

    styles = getSampleStyleSheet()
    h_style = ParagraphStyle("h", parent=styles["Heading1"],
                             fontName="Helvetica-Bold", fontSize=18,
                             textColor=colors.HexColor("#3B1A2A"), spaceAfter=2)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontName="Helvetica", fontSize=8,
                               textColor=colors.HexColor("#6B7280"), spaceAfter=10)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"],
                                fontName="Helvetica", fontSize=8.5,
                                textColor=colors.HexColor("#1F2937"), spaceAfter=4)

    story = []
    story.append(Paragraph("KAZO · Audience Report", h_style))
    story.append(Paragraph("Powered by Fundle · Real-time loyalty data", sub_style))
    story.append(Paragraph(f"<b>Segment:</b> {seg_name}", meta_style))
    story.append(Paragraph(
        f"<b>Generated:</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} &nbsp;·&nbsp; "
        f"<b>By:</b> {user.get('email','—')}",
        meta_style,
    ))
    story.append(Paragraph(
        f"<b>Total matched customers:</b> {total_matched:,} &nbsp;·&nbsp; "
        f"<b>Rows in this PDF:</b> {len(rows_collected):,} of {total_exported:,} exported",
        meta_style,
    ))
    if truncated_at:
        story.append(Paragraph(
            f"<i>PDF table truncated to first {PDF_TABLE_CAP:,} rows for readability. "
            f"Use CSV / XLSX for the full {total_exported:,}-row dataset.</i>",
            sub_style,
        ))
    story.append(Spacer(1, 6))

    header_row = [label for _, label, _ in pdf_cols]
    table_data = [header_row] + rows_collected if rows_collected else [header_row, ["No customers matched the filter."] + [""] * (len(pdf_cols) - 1)]
    col_widths = [w * mm for _, _, w in pdf_cols]
    tbl = Table(table_data, repeatRows=1, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B1A2A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
        ("ALIGN",      (4, 1), (4, -1), "RIGHT"),
        ("ALIGN",      (5, 1), (5, -1), "RIGHT"),
        ("ALIGN",      (7, 1), (7, -1), "RIGHT"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF7F4")]),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#9CA3AF"))
        canvas.drawString(
            14 * mm, 8 * mm,
            f"KAZO · Fundle · {seg_name} · Page {doc_.page}",
        )
        canvas.drawRightString(
            doc_.pagesize[0] - 14 * mm, 8 * mm,
            "Confidential — internal use only",
        )
        canvas.restoreState()

    doc_pdf.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    filename = f"{safe_name}_{ts}.pdf"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/")
async def create_segment(body: SegmentIn, user: dict = Depends(get_current_user)):
    """Save a named segment. Returns the saved doc with cached counts."""
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "Segment name is required")
    # Pre-compute counts for the list view
    pre = await _compute_preview(body.tree, body.window)
    doc = {
        "id": uuid.uuid4().hex,
        "name": name,
        "description": (body.description or "").strip(),
        "tree": body.tree,
        "window": body.window,
        "matched_total": pre["matched_total"],
        "reach": pre["reach"],
        "created_at": _iso(_now()),
        "created_by": user["email"],
        "created_by_name": user.get("name"),
        "updated_at": _iso(_now()),
    }
    await segments_col.insert_one({**doc})  # _id is auto, we exclude it on read
    return doc


@router.get("/")
async def list_segments(user: dict = Depends(get_current_user)):
    rows = await segments_col.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"rows": rows}


@router.get("/{segment_id}")
async def get_segment(segment_id: str, user: dict = Depends(get_current_user)):
    d = await segments_col.find_one({"id": segment_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Segment not found")
    return d


@router.put("/{segment_id}")
async def update_segment(segment_id: str, body: SegmentIn, user: dict = Depends(get_current_user)):
    existing = await segments_col.find_one({"id": segment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Segment not found")
    if existing["created_by"] != user["email"] and user.get("role") not in {"super_admin", "brand_admin"}:
        raise HTTPException(403, "Only the creator or an admin can edit this segment")
    pre = await _compute_preview(body.tree, body.window)
    update = {
        "name": body.name.strip(),
        "description": (body.description or "").strip(),
        "tree": body.tree,
        "window": body.window,
        "matched_total": pre["matched_total"],
        "reach": pre["reach"],
        "updated_at": _iso(_now()),
    }
    await segments_col.update_one({"id": segment_id}, {"$set": update})
    refreshed = await segments_col.find_one({"id": segment_id}, {"_id": 0})
    return refreshed


@router.delete("/{segment_id}")
async def delete_segment(segment_id: str, user: dict = Depends(get_current_user)):
    existing = await segments_col.find_one({"id": segment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Segment not found")
    if existing["created_by"] != user["email"] and user.get("role") not in {"super_admin", "brand_admin"}:
        raise HTTPException(403, "Only the creator or an admin can delete this segment")
    await segments_col.delete_one({"id": segment_id})
    return {"deleted": True}


@router.post("/{segment_id}/refresh")
async def refresh_segment(segment_id: str, user: dict = Depends(get_current_user)):
    s = await segments_col.find_one({"id": segment_id}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Segment not found")
    pre = await _compute_preview(s["tree"], s.get("window"))
    await segments_col.update_one({"id": segment_id},
                                    {"$set": {"matched_total": pre["matched_total"],
                                              "reach": pre["reach"],
                                              "updated_at": _iso(_now())}})
    return {"matched_total": pre["matched_total"], "reach": pre["reach"]}
