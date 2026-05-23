"""Raw Data Reports — high-density operational reports.

Five reports modelled after the screenshots provided by the brand:
    1. Customer Data         — customers grouped by Location / City / State / Zone / Month / Tier
    2. Transaction Data      — txn rollup (customers, bills, purchase, points) grouped as above
    3. Repeat Purchases      — Purchase + Repeat (Total / Current / Earlier) split per location
    4. Earn-Redeem           — Total Earn / Redeem / Bonus / Expired + Liability per location
    5. Customers by Visit    — frequency distribution: how many customers visited N times

Every row is drill-down clickable: the response shape includes a `drill_key`
that the frontend uses to call /audience to list the underlying customers.

All endpoints respect R1 (bill_date is source of truth for time) and R5
(loyalty data only — bills without customer_mobile are excluded).
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
import uuid
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from auth import get_current_user
from database import customers_col, transactions_col, stores_col, db
from routes._loyalty import LOYALTY_TX_MATCH

logger = logging.getLogger("kazo-fundle.raw_reports")

router = APIRouter(prefix="/raw-reports", tags=["raw-reports"])

# Frequency / current vs earlier cut-off (90 days). Matches eWards convention
# where Current = last 90 days, Earlier = before that.
CURRENT_WINDOW_DAYS = 90


# ============================================================
# Common request models
# ============================================================
class ReportFilter(BaseModel):
    start_date: Optional[str] = None  # YYYY-MM-DD
    end_date: Optional[str] = None    # YYYY-MM-DD
    group_by: str = "location"        # location | city | state | zone | month | tier
    location: Optional[str] = None    # extra refine (store name)
    tier: Optional[str] = None        # extra refine (silver / gold / etc.)
    search: Optional[str] = None      # free-text contained in group_key
    sort_by: Optional[str] = None
    sort_dir: int = -1
    page: int = 1
    page_size: int = 200              # generous default — these are aggregate rows


# Helper expression: convert any bill_date (string OR datetime) to a YYYY-MM string.
# Bills stored from CSV ingest are ISO strings; bills from POS APIs are BSON datetimes.
_MONTH_KEY_TXN = {
    "$cond": {
        "if":   {"$eq": [{"$type": "$bill_date"}, "string"]},
        "then": {"$substr": ["$bill_date", 0, 7]},
        "else": {"$dateToString": {"format": "%Y-%m", "date": "$bill_date"}},
    }
}

_MONTH_KEY_CUST_FIRST = {
    "$cond": {
        "if":   {"$eq": [{"$type": "$first_purchase_at"}, "string"]},
        "then": {"$substr": ["$first_purchase_at", 0, 7]},
        "else": {"$dateToString": {"format": "%Y-%m",
                                       "date": {"$ifNull": ["$first_purchase_at", None]}}},
    }
}


# Group key column resolution (txn-level)
GROUP_FIELD_TXN = {
    "location": "$store_name",
    "city":     "$city",
    "state":    {"$ifNull": ["$state", "$zone"]},
    "zone":     "$zone",
    "month":    _MONTH_KEY_TXN,
    "tier":     None,  # resolved via customer join — special-case
}

# Group key column resolution (customer-level)
GROUP_FIELD_CUST = {
    "location": None,   # resolved via customer.home_store_id → store_name
    "city":     "$city",
    "state":    "$state",
    "zone":     None,   # resolved via home store
    "month":    None,   # not meaningful for static customer base — falls back to first_purchase month
    "tier":     "$tier",
}


def _parse_iso(d: Optional[str]) -> Optional[datetime]:
    if not d:
        return None
    try:
        return datetime.strptime(d, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _date_match(f: ReportFilter) -> Dict[str, Any]:
    """Build a {bill_date: {$gte, $lte}} clause from filter dates.

    Note: bill_date is stored as ISO string in CSV-ingested data and as BSON
    datetime in POS-ingested data. MongoDB's $gte / $lte work natively for
    ISO-8601 strings (lexicographic order matches chronological order) AND
    for datetimes, so the same clause works for both.
    """
    s = _parse_iso(f.start_date)
    e = _parse_iso(f.end_date)
    if not s and not e:
        return {}
    clause: Dict[str, Any] = {}
    if s:
        clause["$gte"] = s
    if e:
        # Include the entire end day
        clause["$lte"] = e + timedelta(days=1) - timedelta(seconds=1)
    # When bill_date is stored as a string, MongoDB's comparison still works
    # if both operands are comparable. Cast datetime → string for safety when
    # bill_date is unmixed type. We use $expr with $convert to be safe.
    return {"$or": [
        {"bill_date": clause},
        {"bill_date": {k: v.isoformat() for k, v in clause.items()}},
    ]}


def _base_match(f: ReportFilter) -> Dict[str, Any]:
    m = {**LOYALTY_TX_MATCH, **_date_match(f)}
    if f.location:
        m["store_name"] = f.location
    return m


def _apply_sort_and_paginate(rows: List[Dict[str, Any]], f: ReportFilter,
                                default_key: str = "total_customers") -> Tuple[List[Dict[str, Any]], int]:
    """Sort + paginate the aggregate rows in memory (counts of rows are tiny — at most a few hundred)."""
    if f.search:
        needle = f.search.lower()
        rows = [r for r in rows if needle in str(r.get("group_key", "")).lower()]
    key = f.sort_by or default_key
    rows = sorted(rows, key=lambda r: (r.get(key) is None, r.get(key)), reverse=(f.sort_dir == -1))
    total = len(rows)
    start = max(0, (f.page - 1) * f.page_size)
    end = start + f.page_size
    return rows[start:end], total


# ============================================================
# 1) Customer Data
# ============================================================
@router.post("/customer-data")
async def customer_data(f: ReportFilter, user: dict = Depends(get_current_user)):
    """Customer counts grouped by Location / City / State / Zone / Month / Tier.

    For Location/City/State/Zone — count distinct customer_mobile from txns in
    the date range (a customer "belongs to" a location if they have at least
    one bill there). For Tier/Month — group customers themselves.
    """
    gb = f.group_by
    if gb not in {"location", "city", "state", "zone", "month", "tier"}:
        raise HTTPException(400, "group_by must be one of location/city/state/zone/month/tier")

    if gb == "tier":
        # Group customers by tier — independent of date range
        agg = customers_col.aggregate([
            {"$match": {"mobile": {"$nin": [None, ""]}}},
            {"$group": {
                "_id": "$tier",
                "total_customers": {"$sum": 1},
                "total_lifetime_spend": {"$sum": {"$ifNull": ["$lifetime_spend", 0]}},
                "total_lifetime_points_earned": {"$sum": {"$ifNull": ["$lifetime_points_earned", 0]}},
                "total_points_balance": {"$sum": {"$ifNull": ["$points_balance", 0]}},
                "total_visit_count": {"$sum": {"$ifNull": ["$visit_count", 0]}},
            }},
            {"$project": {
                "_id": 0, "group_key": "$_id",
                "total_customers": 1,
                "total_lifetime_spend": {"$round": ["$total_lifetime_spend", 2]},
                "total_lifetime_points_earned": 1,
                "total_points_balance": 1,
                "avg_lifetime_spend": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                                 {"$divide": ["$total_lifetime_spend", "$total_customers"]}, 0]}, 2]},
                "avg_visit_count": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                              {"$divide": ["$total_visit_count", "$total_customers"]}, 0]}, 2]},
            }},
        ])
    elif gb == "month":
        # Customers' first_purchase_at month — handle string and datetime types
        agg = customers_col.aggregate([
            {"$match": {"mobile": {"$nin": [None, ""]}, "first_purchase_at": {"$ne": None}}},
            {"$group": {
                "_id": _MONTH_KEY_CUST_FIRST,
                "total_customers": {"$sum": 1},
                "total_lifetime_spend": {"$sum": {"$ifNull": ["$lifetime_spend", 0]}},
                "total_lifetime_points_earned": {"$sum": {"$ifNull": ["$lifetime_points_earned", 0]}},
                "total_points_balance": {"$sum": {"$ifNull": ["$points_balance", 0]}},
                "total_visit_count": {"$sum": {"$ifNull": ["$visit_count", 0]}},
            }},
            {"$project": {
                "_id": 0, "group_key": "$_id",
                "total_customers": 1,
                "total_lifetime_spend": {"$round": ["$total_lifetime_spend", 2]},
                "total_lifetime_points_earned": 1,
                "total_points_balance": 1,
                "avg_lifetime_spend": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                                 {"$divide": ["$total_lifetime_spend", "$total_customers"]}, 0]}, 2]},
                "avg_visit_count": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                              {"$divide": ["$total_visit_count", "$total_customers"]}, 0]}, 2]},
            }},
        ])
    else:
        # location / city / state / zone — group transactions by store/city/state/zone,
        # then enrich with per-group customer roll-ups
        field = GROUP_FIELD_TXN[gb]
        agg = transactions_col.aggregate([
            {"$match": _base_match(f)},
            # Stage 1: per-(group, mobile) per-bill snapshot
            {"$group": {
                "_id": {"key": field, "mobile": "$customer_mobile", "bill": "$bill_number"},
                "net": {"$sum": "$net_amount"},
                "earn": {"$sum": "$points_earned"},
            }},
            # Stage 2: per-(group, mobile): bills count + spend + earn
            {"$group": {
                "_id": {"key": "$_id.key", "mobile": "$_id.mobile"},
                "bills": {"$sum": 1},
                "spend": {"$sum": "$net"},
                "earn":  {"$sum": "$earn"},
            }},
            # Stage 3: per-group rollup with repeat detection
            {"$group": {
                "_id": "$_id.key",
                "total_customers":  {"$sum": 1},
                "total_bills":      {"$sum": "$bills"},
                "total_purchase":   {"$sum": "$spend"},
                "total_earn_points": {"$sum": "$earn"},
                "repeat_customers": {"$sum": {"$cond": [{"$gt": ["$bills", 1]}, 1, 0]}},
                "one_timer_customers": {"$sum": {"$cond": [{"$eq": ["$bills", 1]}, 1, 0]}},
            }},
            {"$project": {
                "_id": 0, "group_key": "$_id",
                "total_customers": 1,
                "total_bills": 1,
                "total_purchase":     {"$round": ["$total_purchase", 2]},
                "total_earn_points":  {"$round": ["$total_earn_points", 2]},
                "repeat_customers": 1,
                "one_timer_customers": 1,
                "avg_lifetime_spend": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                                 {"$divide": ["$total_purchase", "$total_customers"]}, 0]}, 2]},
                "avg_bills_per_customer": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                                      {"$divide": ["$total_bills", "$total_customers"]}, 0]}, 2]},
                "repeat_pct": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                          {"$multiply": [{"$divide": ["$repeat_customers", "$total_customers"]}, 100]}, 0]}, 1]},
            }},
        ])

    raw = [r async for r in agg]
    raw = [r for r in raw if r.get("group_key") not in (None, "")]
    rows, total = _apply_sort_and_paginate(raw, f, default_key="total_customers")
    grand_total = sum(r["total_customers"] for r in raw)
    totals: Dict[str, Any] = {"total_customers": grand_total}
    # Aggregate optional metrics into totals row
    for fld in ("total_bills", "total_purchase", "total_earn_points", "repeat_customers",
                 "one_timer_customers", "total_lifetime_spend", "total_lifetime_points_earned",
                 "total_points_balance"):
        if raw and fld in raw[0]:
            totals[fld] = round(sum(r.get(fld, 0) for r in raw), 2)
    if "total_purchase" in totals and grand_total:
        totals["avg_lifetime_spend"] = round(totals["total_purchase"] / grand_total, 2)
    if "total_bills" in totals and grand_total:
        totals["avg_bills_per_customer"] = round(totals["total_bills"] / grand_total, 2)
    if "repeat_customers" in totals and grand_total:
        totals["repeat_pct"] = round(totals["repeat_customers"] * 100 / grand_total, 1)
    return {
        "group_by": gb,
        "rows": rows,
        "total": total,
        "totals": totals,
        "chart": [{"label": r["group_key"], "value": r["total_customers"]} for r in raw[:30]],
    }


# ============================================================
# 2) Transaction Data
# ============================================================
@router.post("/transaction-data")
async def transaction_data(f: ReportFilter, user: dict = Depends(get_current_user)):
    """Per-group rollup: Total Customers, Total Bills, Total Purchase, Total Earn Points."""
    gb = f.group_by
    if gb not in {"location", "city", "state", "zone", "month"}:
        raise HTTPException(400, "group_by must be one of location/city/state/zone/month")
    field = GROUP_FIELD_TXN[gb]

    agg = transactions_col.aggregate([
        {"$match": _base_match(f)},
        # Stage 1: build per-(group,mobile) per-bill snapshot
        {"$group": {
            "_id": {"key": field, "mobile": "$customer_mobile", "bill": "$bill_number"},
            "net":      {"$sum": "$net_amount"},
            "gross":    {"$sum": {"$ifNull": ["$gross_amount", "$net_amount"]}},
            "discount": {"$sum": {"$ifNull": ["$discount_amount", 0]}},
            "earn":     {"$sum": "$points_earned"},
        }},
        # Stage 2: roll up to group level
        {"$group": {
            "_id": "$_id.key",
            "total_customers": {"$addToSet": "$_id.mobile"},
            "total_bills": {"$sum": 1},
            "total_purchase": {"$sum": "$net"},
            "total_gross_purchase": {"$sum": "$gross"},
            "total_discount": {"$sum": "$discount"},
            "total_earn_points": {"$sum": "$earn"},
        }},
        {"$project": {
            "_id": 0,
            "group_key": "$_id",
            "total_customers": {"$size": "$total_customers"},
            "total_bills": 1,
            "total_purchase":      {"$round": ["$total_purchase", 2]},
            "total_gross_purchase": {"$round": ["$total_gross_purchase", 2]},
            "total_discount":      {"$round": ["$total_discount", 2]},
            "total_earn_points":   {"$round": ["$total_earn_points", 2]},
            "avg_bill_value": {"$round": [{"$cond": [{"$gt": ["$total_bills", 0]},
                                                          {"$divide": ["$total_purchase", "$total_bills"]}, 0]}, 2]},
            "avg_customer_spend": {"$round": [{"$cond": [{"$gt": [{"$size": "$total_customers"}, 0]},
                                                              {"$divide": ["$total_purchase", {"$size": "$total_customers"}]}, 0]}, 2]},
            "discount_pct": {"$round": [{"$cond": [{"$gt": ["$total_gross_purchase", 0]},
                                                       {"$multiply": [{"$divide": ["$total_discount", "$total_gross_purchase"]}, 100]}, 0]}, 1]},
        }},
    ])
    raw = [r async for r in agg if r.get("group_key") not in (None, "")]
    rows, total = _apply_sort_and_paginate(raw, f, default_key="total_purchase")
    totals = {
        "total_customers": sum(r["total_customers"] for r in raw),
        "total_bills": sum(r["total_bills"] for r in raw),
        "total_purchase": round(sum(r["total_purchase"] for r in raw), 2),
        "total_gross_purchase": round(sum(r.get("total_gross_purchase", 0) for r in raw), 2),
        "total_discount": round(sum(r.get("total_discount", 0) for r in raw), 2),
        "total_earn_points": round(sum(r["total_earn_points"] for r in raw), 2),
    }
    if totals["total_bills"]:
        totals["avg_bill_value"] = round(totals["total_purchase"] / totals["total_bills"], 2)
    if totals["total_customers"]:
        totals["avg_customer_spend"] = round(totals["total_purchase"] / totals["total_customers"], 2)
    if totals["total_gross_purchase"]:
        totals["discount_pct"] = round(totals["total_discount"] * 100 / totals["total_gross_purchase"], 1)
    return {
        "group_by": gb,
        "rows": rows,
        "total": total,
        "totals": totals,
        "chart": [
            {
                "label": r["group_key"],
                "total_purchase": r["total_purchase"],
                "total_bills": r["total_bills"],
                "total_earn_points": r["total_earn_points"],
                "total_customers": r["total_customers"],
            } for r in raw[:30]
        ],
    }


# ============================================================
# 3) Repeat Purchases
# ============================================================
@router.post("/repeat-purchases")
async def repeat_purchases(f: ReportFilter, user: dict = Depends(get_current_user)):
    """Per-location split: Purchase totals + Repeat Purchase (Total/Current/Earlier).

    Definitions:
      - Purchase   = ALL loyalty bills in the date range (or all-time if no range)
      - Repeat     = bills from customers who already had ≥1 prior bill at the
                      same store before this one (i.e. 2nd+ visit)
      - Current    = repeat bills within the last 90 days
      - Earlier    = repeat bills older than 90 days but within the date range
    """
    gb = f.group_by
    if gb not in {"location", "city", "state", "zone", "month"}:
        raise HTTPException(400, "group_by must be one of location/city/state/zone/month")
    field = GROUP_FIELD_TXN[gb]
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=CURRENT_WINDOW_DAYS)

    # 1) Per-mobile-per-store first-bill timestamp — to know which bills are "repeats"
    pipeline = [
        {"$match": _base_match(f)},
        {"$sort": {"bill_date": 1}},
        {"$group": {
            "_id": {"mobile": "$customer_mobile", "key": field},
            "bills": {"$push": {"net": "$net_amount", "date": "$bill_date",
                                  "bill_no": "$bill_number"}},
        }},
        # Per-mobile-key compute totals + repeat split in JS-side projection
    ]
    cursor = transactions_col.aggregate(pipeline, allowDiskUse=True)

    def _to_dt(v: Any) -> Optional[datetime]:
        if isinstance(v, datetime):
            return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
        if isinstance(v, str):
            try:
                # Handle "2026-05-01T..." and "2026-05-01"
                s = v.replace("Z", "+00:00")
                d = datetime.fromisoformat(s)
                return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
            except Exception:
                return None
        return None

    # Aggregate in Python — for very large datasets we'd push to $reduce, but
    # mobile×location pairs are bounded by customer count.
    agg_map: Dict[str, Dict[str, Any]] = {}
    async for doc in cursor:
        _id = doc.get("_id") or {}
        if not isinstance(_id, dict):
            continue
        mobile = _id.get("mobile")
        key = _id.get("key")
        if not mobile or not key:
            continue
        bills = doc.get("bills", [])
        for b in bills:
            b["_dt"] = _to_dt(b.get("date"))
        bills.sort(key=lambda b: b.get("_dt") or datetime.min.replace(tzinfo=timezone.utc))
        unique_bill_nos = {}
        for b in bills:
            bn = b.get("bill_no") or f"unk-{id(b)}"
            if bn not in unique_bill_nos:
                unique_bill_nos[bn] = b
        bills_unique = list(unique_bill_nos.values())
        bills_unique.sort(key=lambda b: b.get("_dt") or datetime.min.replace(tzinfo=timezone.utc))

        row = agg_map.setdefault(key, {
            "group_key": key,
            "purchase_unique_customers": set(),
            "purchase_total_bills": 0,
            "purchase_total_purchase": 0.0,
            "repeat_total_unique_customers": set(),
            "repeat_total_bills": 0,
            "repeat_total_purchase": 0.0,
            "repeat_current_unique_customers": set(),
            "repeat_current_bills": 0,
            "repeat_current_purchase": 0.0,
            "repeat_earlier_unique_customers": set(),
            "repeat_earlier_bills": 0,
            "repeat_earlier_purchase": 0.0,
        })
        row["purchase_unique_customers"].add(mobile)
        row["purchase_total_bills"] += len(bills_unique)
        row["purchase_total_purchase"] += sum(b.get("net") or 0 for b in bills_unique)

        if len(bills_unique) > 1:
            row["repeat_total_unique_customers"].add(mobile)
            row["repeat_total_bills"] += len(bills_unique) - 1
            row["repeat_total_purchase"] += sum(b.get("net") or 0 for b in bills_unique[1:])

            for b in bills_unique[1:]:
                d = b.get("_dt")
                net = b.get("net") or 0
                if d and d >= cutoff:
                    row["repeat_current_unique_customers"].add(mobile)
                    row["repeat_current_bills"] += 1
                    row["repeat_current_purchase"] += net
                else:
                    row["repeat_earlier_unique_customers"].add(mobile)
                    row["repeat_earlier_bills"] += 1
                    row["repeat_earlier_purchase"] += net

    raw: List[Dict[str, Any]] = []
    for k, row in agg_map.items():
        raw.append({
            "group_key": k,
            "purchase_unique_customers": len(row["purchase_unique_customers"]),
            "purchase_total_bills": row["purchase_total_bills"],
            "purchase_total_purchase": round(row["purchase_total_purchase"], 2),
            "repeat_total_unique_customers": len(row["repeat_total_unique_customers"]),
            "repeat_total_bills": row["repeat_total_bills"],
            "repeat_total_purchase": round(row["repeat_total_purchase"], 2),
            "repeat_current_unique_customers": len(row["repeat_current_unique_customers"]),
            "repeat_current_bills": row["repeat_current_bills"],
            "repeat_current_purchase": round(row["repeat_current_purchase"], 2),
            "repeat_earlier_unique_customers": len(row["repeat_earlier_unique_customers"]),
            "repeat_earlier_bills": row["repeat_earlier_bills"],
            "repeat_earlier_purchase": round(row["repeat_earlier_purchase"], 2),
        })

    rows, total = _apply_sort_and_paginate(raw, f, default_key="purchase_total_purchase")
    totals: Dict[str, Any] = {}
    for fld in ("purchase_unique_customers", "purchase_total_bills", "purchase_total_purchase",
                "repeat_total_unique_customers", "repeat_total_bills", "repeat_total_purchase",
                "repeat_current_unique_customers", "repeat_current_bills", "repeat_current_purchase",
                "repeat_earlier_unique_customers", "repeat_earlier_bills", "repeat_earlier_purchase"):
        s = sum(r[fld] for r in raw)
        totals[fld] = round(s, 2) if "purchase" in fld else s

    return {
        "group_by": gb,
        "rows": rows,
        "total": total,
        "totals": totals,
        "current_window_days": CURRENT_WINDOW_DAYS,
    }


# ============================================================
# 4) Earn Redeem
# ============================================================
@router.post("/earn-redeem")
async def earn_redeem(f: ReportFilter, user: dict = Depends(get_current_user)):
    """Per-group points rollup. Liability = Earn + Bonus - Redeem - Expired."""
    gb = f.group_by
    if gb not in {"location", "city", "state", "zone", "month"}:
        raise HTTPException(400, "group_by must be one of location/city/state/zone/month")
    field = GROUP_FIELD_TXN[gb]

    # Earn / Redeem / Bonus from transactions
    agg = transactions_col.aggregate([
        {"$match": _base_match(f)},
        {"$group": {
            "_id": field,
            "total_earn_points": {"$sum": "$points_earned"},
            "total_redeem_points": {"$sum": "$points_redeemed"},
            "total_bonus_points": {"$sum": {"$ifNull": ["$bonus_points", 0]}},
        }},
        {"$project": {
            "_id": 0,
            "group_key": "$_id",
            "total_earn_points":   {"$round": ["$total_earn_points", 2]},
            "total_redeem_points": {"$round": ["$total_redeem_points", 2]},
            "total_bonus_points":  {"$round": ["$total_bonus_points", 2]},
        }},
    ])
    txn_rows = {r["group_key"]: r async for r in agg if r.get("group_key") not in (None, "")}

    # Expired points from points_ledger (type=expire entries)
    plc = db["points_ledger"]
    expired_agg = plc.aggregate([
        {"$match": {"type": "expire",
                    **({"created_at": {"$gte": _parse_iso(f.start_date),
                                        "$lte": _parse_iso(f.end_date) + timedelta(days=1) - timedelta(seconds=1)
                                            if _parse_iso(f.end_date) else None}}
                       if (f.start_date or f.end_date) else {})}},
        {"$group": {"_id": None, "total": {"$sum": {"$abs": "$points"}}}},
    ])
    expired_total = 0
    async for r in expired_agg:
        expired_total = abs(r.get("total") or 0)

    # Distribute expired proportionally to redeem totals — without per-bill linkage
    # we approximate; for now show as a separate aggregate column ("global expired")
    raw: List[Dict[str, Any]] = []
    total_redeem_global = sum(v["total_redeem_points"] for v in txn_rows.values()) or 1
    for key, r in txn_rows.items():
        # Pro-rate expired by group's share of redemption (best-effort approximation)
        prorated_expired = round(expired_total * (r["total_redeem_points"] / total_redeem_global), 2) \
                              if expired_total else 0
        liability = round(
            r["total_earn_points"] + r["total_bonus_points"]
            - r["total_redeem_points"] - prorated_expired,
            2,
        )
        gross_earned = r["total_earn_points"] + r["total_bonus_points"]
        redemption_rate = round((r["total_redeem_points"] / gross_earned * 100), 1) if gross_earned else 0.0
        raw.append({
            **r,
            "total_expired_points": prorated_expired,
            "total_liability": liability,
            "gross_points_earned": round(gross_earned, 2),
            "redemption_rate_pct": redemption_rate,
        })

    rows, total = _apply_sort_and_paginate(raw, f, default_key="total_earn_points")
    totals = {
        "total_earn_points":   round(sum(r["total_earn_points"] for r in raw), 2),
        "total_redeem_points": round(sum(r["total_redeem_points"] for r in raw), 2),
        "total_bonus_points":  round(sum(r["total_bonus_points"] for r in raw), 2),
        "total_expired_points": round(sum(r["total_expired_points"] for r in raw), 2),
        "total_liability":     round(sum(r["total_liability"] for r in raw), 2),
        "gross_points_earned": round(sum(r["gross_points_earned"] for r in raw), 2),
    }
    if totals["gross_points_earned"]:
        totals["redemption_rate_pct"] = round(
            totals["total_redeem_points"] * 100 / totals["gross_points_earned"], 1
        )
    return {
        "group_by": gb,
        "rows": rows,
        "total": total,
        "totals": totals,
        "chart": [
            {"label": r["group_key"], "earn": r["total_earn_points"],
             "redeem": r["total_redeem_points"], "bonus": r["total_bonus_points"],
             "expired": r["total_expired_points"]}
            for r in raw[:30]
        ],
    }


# ============================================================
# 5) Customers by Visit
# ============================================================
@router.post("/customers-by-visit")
async def customers_by_visit(f: ReportFilter, user: dict = Depends(get_current_user)):
    """Frequency distribution of unique customers by their visit count in the
    given window. Visits = unique bill_number in the window."""
    pipeline: List[Dict[str, Any]] = [
        {"$match": _base_match(f)},
        {"$group": {
            "_id": {"mobile": "$customer_mobile", "bill": "$bill_number"},
            "net": {"$sum": "$net_amount"},
        }},
        {"$group": {
            "_id": "$_id.mobile",
            "visits": {"$sum": 1},
            "spend":  {"$sum": "$net"},
        }},
        {"$group": {
            "_id": "$visits",
            "total_customers": {"$sum": 1},
            "total_spend": {"$sum": "$spend"},
        }},
        {"$project": {
            "_id": 0,
            "visits": "$_id",
            "total_customers": 1,
            "total_purchase": {"$round": ["$total_spend", 2]},
            "avg_customer_spend": {"$round": [{"$cond": [{"$gt": ["$total_customers", 0]},
                                                              {"$divide": ["$total_spend", "$total_customers"]}, 0]}, 2]},
        }},
        {"$sort": {"visits": 1}},
    ]

    # Refine by location/tier if requested
    if f.location or f.tier:
        match: Dict[str, Any] = _base_match(f)
        if f.location:
            match["store_name"] = f.location
        if f.tier:
            tier_mobiles = await customers_col.distinct("mobile", {"tier": f.tier})
            match["customer_mobile"] = {"$in": tier_mobiles}
        pipeline[0] = {"$match": match}

    raw = [r async for r in transactions_col.aggregate(pipeline, allowDiskUse=True)]
    grand_visits = sum(r["visits"] * r["total_customers"] for r in raw)
    grand_customers = sum(r["total_customers"] for r in raw)
    grand_purchase = round(sum(r.get("total_purchase", 0) for r in raw), 2)
    return {
        "rows": raw,
        "totals": {
            "visits": grand_visits,
            "total_customers": grand_customers,
            "total_purchase": grand_purchase,
            "avg_customer_spend": round(grand_purchase / grand_customers, 2) if grand_customers else 0,
        },
    }


# ============================================================
# Aggregate drill-down — return underlying customer list for a cell
# ============================================================
class DrillIn(BaseModel):
    report: str          # customer-data | transaction-data | repeat-purchases | earn-redeem | customers-by-visit
    group_by: str
    group_key: str       # the value of the group cell user clicked
    metric: Optional[str] = None  # purchase_total_bills / repeat_total_purchase / etc.
    visits: Optional[int] = None  # for customers-by-visit
    filters: ReportFilter
    page: int = 1
    page_size: int = 50


@router.post("/drill")
async def drill(body: DrillIn, user: dict = Depends(get_current_user)):
    """Return the underlying customer list for any cell in any report."""
    f = body.filters
    match = _base_match(f)
    if body.group_by == "location":
        match["store_name"] = body.group_key
    elif body.group_by == "city":
        match["city"] = body.group_key
    elif body.group_by == "zone":
        match["zone"] = body.group_key
    elif body.group_by == "state":
        # state is stored as zone for txns; or could be city — try both
        match["$or"] = [{"state": body.group_key}, {"zone": body.group_key}]
    elif body.group_by == "month":
        # Build a bill_date range for the YYYY-MM
        try:
            yr, mo = body.group_key.split("-")
            yr, mo = int(yr), int(mo)
            start = datetime(yr, mo, 1, tzinfo=timezone.utc)
            end = (datetime(yr + (1 if mo == 12 else 0),
                              1 if mo == 12 else mo + 1, 1, tzinfo=timezone.utc)
                     - timedelta(seconds=1))
            match["bill_date"] = {"$gte": start, "$lte": end}
        except Exception:
            pass

    if body.visits is not None:
        # Restrict to customers who had exactly this many bills
        agg = transactions_col.aggregate([
            {"$match": match},
            {"$group": {"_id": {"m": "$customer_mobile", "b": "$bill_number"}}},
            {"$group": {"_id": "$_id.m", "visits": {"$sum": 1}}},
            {"$match": {"visits": body.visits}},
        ])
        mobiles = [r["_id"] async for r in agg]
    else:
        mobiles = await transactions_col.distinct("customer_mobile", match)

    total = len(mobiles)
    start = max(0, (body.page - 1) * body.page_size)
    end = start + body.page_size
    page_mobiles = mobiles[start:end]

    customers = await customers_col.find(
        {"mobile": {"$in": page_mobiles}},
        {"_id": 0, "mobile": 1, "name": 1, "email": 1, "city": 1, "tier": 1,
         "visit_count": 1, "lifetime_spend": 1, "last_visit_at": 1,
         "points_balance": 1},
    ).to_list(len(page_mobiles))

    return {
        "total": total,
        "page": body.page,
        "page_size": body.page_size,
        "pages": (total + body.page_size - 1) // body.page_size if body.page_size else 1,
        "rows": customers,
    }


# ============================================================
# AI Narrative — Fundle Brain commentary for any report
# ============================================================
class NarrativeIn(BaseModel):
    report: str
    group_by: str
    rows: List[Dict[str, Any]]
    totals: Dict[str, Any] = Field(default_factory=dict)
    filters: ReportFilter


@router.post("/narrative")
async def narrative(body: NarrativeIn, user: dict = Depends(get_current_user)):
    """Ask Fundle Brain (GPT-5) for a 3-bullet commentary on the report rows."""
    key = os.environ.get("EMERGENT_LLM_KEY")

    # Trim large rows arrays for prompt size — top 25 by first numeric column
    rows = body.rows[:25]
    rows_text = "\n".join(
        "  " + " · ".join(f"{k}={v}" for k, v in r.items() if k != "_id")
        for r in rows
    )

    prompt = (
        f"Loyalty raw-data report: {body.report} (grouped by {body.group_by}).\n"
        f"Date range: {body.filters.start_date or 'all-time'} → {body.filters.end_date or 'now'}.\n"
        f"Top 25 rows:\n{rows_text}\n\n"
        f"Grand totals: {body.totals}\n\n"
        "Write a CRISP 3-bullet executive commentary for the KAZO brand manager. "
        "Each bullet ≤ 25 words. Lead with the most actionable insight. "
        "Use 2 concrete numbers per bullet. Do NOT repeat the totals verbatim — "
        "interpret them. Plain text, no markdown."
    )

    if not key:
        # Template fallback
        bullets = []
        if body.totals:
            top = body.totals
            bullets.append(
                f"Aggregate snapshot — {sum(v for k, v in top.items() if 'customers' in k.lower()):,} "
                f"customers across {len(body.rows)} {body.group_by}s."
            )
        if len(rows) >= 2:
            top_row = rows[0]
            bullets.append(
                f"Top {body.group_by}: {top_row.get('group_key', '—')} leads on the primary metric."
            )
        bullets.append("Drill into the top 3 groups to identify pattern outliers.")
        return {"source": "template_fallback", "bullets": bullets, "narrative": "\n".join(f"• {b}" for b in bullets)}

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage  # type: ignore
        llm = LlmChat(api_key=key, session_id=f"rawreport-{datetime.now(timezone.utc).timestamp()}",
                       system_message=(
                           "You are Fundle Brain, KAZO's loyalty analyst. "
                           "Write 3 short executive bullets about a raw-data report."
                       )).with_model("openai", "gpt-5")
        resp = await llm.send_message(UserMessage(text=prompt))
        text = (resp or "").strip()
        # Split into bullets — pre-existing • or - or numbered prefixes
        bullets = []
        for line in text.splitlines():
            line = line.strip()
            line = re.sub(r"^([•\-\*\d]+[\.\)\s]+)", "", line).strip()
            if line:
                bullets.append(line)
        bullets = bullets[:5]
        return {"source": "fundle_brain_gpt5", "bullets": bullets, "narrative": text}
    except Exception as e:
        logger.warning(f"Narrative failed, falling back: {e}")
        return {"source": "template_fallback", "bullets": [str(e)], "narrative": str(e)}


# ============================================================
# Universal export — same payload, CSV / XLSX / PDF
# ============================================================
class ExportIn(BaseModel):
    report: str                          # display name
    group_by: str
    columns: List[Dict[str, str]]        # [{key, label}, ...]
    rows: List[Dict[str, Any]]
    totals: Dict[str, Any] = Field(default_factory=dict)
    chart_title: Optional[str] = None
    format: str = "csv"                  # csv | xlsx | pdf


@router.post("/export")
async def export_report(body: ExportIn, user: dict = Depends(get_current_user)):
    fmt = (body.format or "csv").lower()
    if fmt not in {"csv", "xlsx", "pdf"}:
        raise HTTPException(400, "format must be csv, xlsx or pdf")

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", body.report)[:60] or "raw_report"
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    cols = body.columns or [{"key": k, "label": k} for k in (body.rows[0].keys() if body.rows else [])]

    if fmt == "csv":
        def _gen():
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow([c["label"] for c in cols])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)
            for r in body.rows:
                w.writerow([r.get(c["key"], "") for c in cols])
                if buf.tell() > 32 * 1024:
                    yield buf.getvalue()
                    buf.seek(0)
                    buf.truncate(0)
            # Totals footer
            if body.totals:
                w.writerow([])
                w.writerow(["TOTALS"] + ["" for _ in cols[1:]])
                for k, v in body.totals.items():
                    w.writerow([k, v])
            if buf.tell() > 0:
                yield buf.getvalue()

        return StreamingResponse(
            _gen(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_{ts}.csv"'},
        )

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = body.report[:30] or "Report"
        # Header
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="3B1A2A")
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=i, value=c["label"])
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for r_idx, r in enumerate(body.rows, start=2):
            for c_idx, c in enumerate(cols, start=1):
                ws.cell(row=r_idx, column=c_idx, value=r.get(c["key"], ""))
        # Auto-width
        for i in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        ws.freeze_panes = "A2"
        # Totals sheet
        if body.totals:
            ms = wb.create_sheet("Totals")
            ms.append(["Metric", "Value"])
            for k, v in body.totals.items():
                ms.append([k, v])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_{ts}.xlsx"'},
        )

    # PDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    pdf_doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"KAZO {body.report}",
    )
    styles = getSampleStyleSheet()
    h_style = ParagraphStyle("h", parent=styles["Heading1"], fontName="Helvetica-Bold",
                              fontSize=18, textColor=colors.HexColor("#3B1A2A"))
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontName="Helvetica",
                                fontSize=8, textColor=colors.HexColor("#6B7280"))
    story = [Paragraph(f"KAZO · {body.report}", h_style),
              Paragraph("Powered by Fundle · Generated " + datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"), sub_style),
              Spacer(1, 6)]

    header_row = [c["label"] for c in cols]
    rows_data = [[r.get(c["key"], "") for c in cols] for r in body.rows[:2000]]
    table_data = [header_row] + rows_data
    if body.totals:
        table_data.append(["TOTAL"] + ["" for _ in cols[1:]])
        table_data.append([list(body.totals.keys())[0]] + [str(v) for v in body.totals.values()])
    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B1A2A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF7F4")]),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
    ]))
    story.append(tbl)

    def _footer(canvas, d):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#9CA3AF"))
        canvas.drawString(14 * mm, 8 * mm, f"KAZO · Fundle · {body.report} · Page {d.page}")
        canvas.restoreState()

    pdf_doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_{ts}.pdf"'},
    )
